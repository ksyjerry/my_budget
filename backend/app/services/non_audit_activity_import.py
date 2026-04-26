"""Parse '비감사 Activity 표준화' Excel file into rows ready for ServiceTaskMaster insert."""
from pathlib import Path
from typing import Optional

import openpyxl


SERVICE_TYPE_SHEET_MAP = {
    "Activity 표준화_회계자문": "AC",
    "Activity 표준화_내부통제": "IC",
    "Activity 표준화_ESG": "ESG",
    "Activity 표준화_Valuation": "VAL",
    "Activity 표준화_통상자문": "TRADE",
    "Activity 표준화_보험계리": "ACT",
    "Activity 표준화_기타비감사": "ETC",
}


def _find_header_row(ws) -> Optional[int]:
    """Find the first row containing the '대분류' keyword — returns row number (1-based)."""
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10), values_only=True), start=1):
        for cell in row:
            if cell and "대분류" in str(cell):
                return idx
    return None


def _column_map(header_row: tuple) -> dict[str, int]:
    """Return {canonical_key: column_index} by matching Korean labels."""
    mapping = {}
    for col_idx, cell in enumerate(header_row):
        if not cell:
            continue
        text = str(cell).strip()
        if "대분류" in text:
            mapping["category"] = col_idx
        elif "중분류" in text:
            mapping["subcategory"] = col_idx
        elif "소분류" in text:
            mapping["detail"] = col_idx
        elif "Budget 관리단위" in text or text == "Budget 관리단위":
            mapping["budget_unit"] = col_idx
        elif text == "비고":
            mapping["role"] = col_idx
    return mapping


def _stripped(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip().replace("\u200b", "")
    return s or None


def parse_non_audit_activities(path: str) -> list[dict]:
    """Return a flat list of dicts ready for ServiceTaskMaster insert."""
    wb = openpyxl.load_workbook(path, data_only=True)
    source_file = Path(path).name
    results: list[dict] = []
    for sheet_name, service_type in SERVICE_TYPE_SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header_idx = _find_header_row(ws)
        if not header_idx:
            continue
        header_row = next(ws.iter_rows(min_row=header_idx, max_row=header_idx, values_only=True))
        cols = _column_map(header_row)
        if "category" not in cols or "detail" not in cols:
            continue
        order = 0
        for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
            category = _stripped(row[cols["category"]]) if cols["category"] < len(row) else None
            subcategory = _stripped(row[cols.get("subcategory", -1)]) if "subcategory" in cols and cols["subcategory"] < len(row) else None
            detail = _stripped(row[cols["detail"]]) if cols["detail"] < len(row) else None
            budget_unit = _stripped(row[cols.get("budget_unit", -1)]) if "budget_unit" in cols and cols["budget_unit"] < len(row) else None
            role = _stripped(row[cols.get("role", -1)]) if "role" in cols and cols["role"] < len(row) else None
            if not category and not detail:
                continue
            order += 1
            results.append({
                "service_type": service_type,
                "task_category": category,
                "activity_subcategory": subcategory,
                "activity_detail": detail,
                "task_name": detail or category or "",
                "budget_unit": budget_unit,
                "role": role,
                "sort_order": order,
                "source_file": source_file,
            })
    wb.close()
    return results


from sqlalchemy.orm import Session as DBSessionType

from app.models.project import ServiceTaskMaster


FEEDBACK_XLSX = Path("/Users/jkim564/Documents/Programming/my_budget/files/Budget+ 2.0 Feedback_0425.xlsx")
FINANCIAL_SHEET = "#04_감사_금융업"


def import_financial_activities(db: DBSessionType) -> dict:
    """Import 금융업 audit activities from #04_감사_금융업 sheet into ServiceTaskMaster.

    POL-03 (a): 소분류명을 별도 컬럼(subcategory_name) 으로 저장.
    Returns count summary.
    """
    if not FEEDBACK_XLSX.exists():
        return {"error": f"feedback xlsx not found at {FEEDBACK_XLSX}"}

    wb = openpyxl.load_workbook(FEEDBACK_XLSX, data_only=True)
    ws = wb[FINANCIAL_SHEET]

    inserted = 0
    skipped = 0
    order = 0

    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or all(v is None for v in row):
            continue
        # Columns: None | cat_code | cat_name | mid_code | mid_name | sub_code | sub_name | budget_unit
        cat_code = _stripped(row[1]) if len(row) > 1 else None
        cat_name = _stripped(row[2]) if len(row) > 2 else None
        mid_code = _stripped(row[3]) if len(row) > 3 else None
        mid_name = _stripped(row[4]) if len(row) > 4 else None
        sub_code = _stripped(row[5]) if len(row) > 5 else None
        sub_name = _stripped(row[6]) if len(row) > 6 else None
        budget_unit = _stripped(row[7]) if len(row) > 7 else None

        if not budget_unit:
            skipped += 1
            continue

        # Idempotent: skip if already exists with same service_type + budget_unit + subcategory_name
        existing = db.query(ServiceTaskMaster).filter(
            ServiceTaskMaster.service_type == "AUDIT",
            ServiceTaskMaster.budget_unit == budget_unit,
            ServiceTaskMaster.subcategory_name == sub_name,
        ).first()
        if existing:
            skipped += 1
            continue

        order += 1
        item = ServiceTaskMaster(
            service_type="AUDIT",
            task_category=cat_name,       # 대분류명 → task_category
            activity_subcategory=mid_name, # 중분류명 → activity_subcategory
            activity_detail=sub_code,     # 소분류코드 → activity_detail
            subcategory_name=sub_name,    # 소분류명 → subcategory_name (POL-03)
            task_name=sub_name or budget_unit,
            budget_unit=budget_unit,
            sort_order=order,
            source_file=FEEDBACK_XLSX.name,
        )
        db.add(item)
        inserted += 1

    db.commit()
    wb.close()
    return {"inserted": inserted, "skipped": skipped}


def import_non_audit_activities(
    db: DBSessionType,
    path: str,
    *,
    truncate: bool = True,
) -> dict:
    """Parse Excel and replace ServiceTaskMaster rows for the 7 non-audit services.

    Returns {"inserted": int, "by_service_type": {code: count, ...}, "source_file": str}.

    AUDIT rows are never touched — only non-audit codes are truncated/re-inserted.
    """
    rows = parse_non_audit_activities(path)
    non_audit_codes = set(SERVICE_TYPE_SHEET_MAP.values())
    if truncate:
        db.query(ServiceTaskMaster).filter(
            ServiceTaskMaster.service_type.in_(non_audit_codes)
        ).delete(synchronize_session=False)
        db.commit()
    by_service_type: dict[str, int] = {code: 0 for code in non_audit_codes}
    for r in rows:
        db.add(ServiceTaskMaster(
            service_type=r["service_type"],
            task_category=r["task_category"],
            task_name=r["task_name"],
            activity_subcategory=r["activity_subcategory"],
            activity_detail=r["activity_detail"],
            budget_unit=r["budget_unit"],
            role=r["role"],
            sort_order=r["sort_order"],
            source_file=r["source_file"],
        ))
        by_service_type[r["service_type"]] += 1
    db.commit()
    return {
        "inserted": sum(by_service_type.values()),
        "by_service_type": by_service_type,
        "source_file": rows[0]["source_file"] if rows else None,
    }
