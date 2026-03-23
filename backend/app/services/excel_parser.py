"""Budget Excel 파일 파서.

개별 프로젝트 Budget 파일과 통합 Budget DB 파일 모두 지원.
"""
import openpyxl
from datetime import datetime
from typing import Optional


def parse_budget_template(file_path: str) -> dict:
    """개별 프로젝트 Budget Excel (D시트 기반) 파싱."""
    wb = openpyxl.load_workbook(file_path, data_only=True)

    result = {
        "client_info": _parse_client_info(wb),
        "project_info": _parse_project_info(wb),
        "team_members": _parse_team_members(wb),
        "budget_details": _parse_budget_details(wb),
    }
    wb.close()
    return result


def _parse_client_info(wb) -> dict:
    ws = wb["B_ET 기본정보"]
    info = {}

    field_map = {
        5: "industry",
        6: "asset_size",
        7: "listing_status",
        8: "business_report",
        9: "gaap",
        10: "consolidated",
        11: "subsidiary_count",
        12: "internal_control",
        13: "initial_audit",
        14: "fiscal_start",
        15: "contract_hours",
    }

    for row_num, field_name in field_map.items():
        val = ws.cell(row=row_num, column=3).value
        info[field_name] = val

    # Project 기본정보
    info["project_code"] = ws.cell(row=18, column=3).value
    info["project_name"] = ws.cell(row=19, column=3).value
    info["department"] = ws.cell(row=20, column=3).value
    info["el_name"] = ws.cell(row=21, column=3).value
    info["el_empno"] = str(ws.cell(row=21, column=5).value or "")
    info["pm_name"] = ws.cell(row=22, column=3).value
    info["pm_empno"] = str(ws.cell(row=22, column=5).value or "")

    # 시간 정보
    info["qrp_hours"] = ws.cell(row=25, column=3).value or 0
    info["rm_hours"] = ws.cell(row=26, column=3).value or 0
    info["el_hours"] = ws.cell(row=27, column=3).value or 0
    info["pm_hours"] = ws.cell(row=28, column=3).value or 0
    info["ra_elpm_hours"] = ws.cell(row=29, column=3).value or 0

    return info


def _parse_project_info(wb) -> dict:
    ws = wb["D_2606 budget template"]
    info = {}
    info["contract_hours"] = ws.cell(row=3, column=3).value or 0
    info["axdx_hours"] = ws.cell(row=4, column=3).value or 0
    info["et_controllable_budget"] = ws.cell(row=9, column=3).value or 0
    info["fulcrum_hours"] = ws.cell(row=10, column=3).value or 0
    info["ra_staff_hours"] = ws.cell(row=11, column=3).value or 0
    info["specialist_hours"] = ws.cell(row=12, column=3).value or 0
    info["travel_hours"] = ws.cell(row=14, column=3).value or 0
    info["template_status"] = ws.cell(row=19, column=3).value
    return info


def _parse_team_members(wb) -> list[dict]:
    ws = wb["C_ET 구성원 정보"]
    members = []
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row):
        role = row[0].value
        name = row[1].value
        empno = row[2].value
        if not role or not name or name.startswith("Defalut") or name.startswith("ET 구성원"):
            continue
        members.append({
            "role": role,
            "name": name,
            "empno": str(empno) if empno else "",
        })
    return members


def _parse_budget_details(wb) -> list[dict]:
    """D시트의 Budget 상세 데이터 파싱."""
    ws = wb["D_2606 budget template"]
    details = []

    # 월 헤더 (K열~V열, 11~22열)
    month_cols = []
    for col in range(11, 23):
        val = ws.cell(row=21, column=col).value
        if isinstance(val, datetime):
            month_cols.append((col, val.strftime("%Y-%m")))
        elif isinstance(val, str) and val.isdigit():
            month_cols.append((col, val))
        else:
            month_cols.append((col, None))

    # 데이터 행 (23행부터)
    for row in ws.iter_rows(min_row=23, max_row=ws.max_row):
        category = row[1].value  # B열: 대분류
        unit = row[2].value      # C열: Budget 관리단위
        is_scope = row[3].value  # D열: 해당 여부
        assignment = row[7].value  # H열: 담당자
        empno = row[8].value     # I열: 사번
        budget_total = row[9].value  # J열: Budget 합계

        if not category or category == "(마지막행)":
            break
        if not is_scope:
            continue

        # 월별 시간
        for col_idx, ym in month_cols:
            hours = row[col_idx - 1].value  # 0-based index
            if hours and hours > 0:
                # year_month 포맷 결정
                if ym and "-" in str(ym):
                    year_month = ym
                elif isinstance(ym, str) and ym.isdigit():
                    month_num = int(ym)
                    year = 2026 if month_num <= 5 else 2025
                    year_month = f"{year}-{month_num:02d}"
                else:
                    continue

                details.append({
                    "budget_category": category,
                    "budget_unit": unit,
                    "emp_name": assignment or "",
                    "empno": str(empno) if empno else "",
                    "year_month": year_month,
                    "budget_hours": float(hours),
                })

    return details


def parse_budget_db_file(file_path: str) -> dict:
    """통합 Budget DB Excel (Budget_데이터_2025.xlsx) 파싱."""
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

    result = {
        "clients": _parse_client_db(wb),
        "projects": _parse_project_db(wb),
        "budget_details": _parse_budget_detail_db(wb),
    }
    wb.close()
    return result


def _parse_client_db(wb) -> list[dict]:
    ws = wb["Client기본정보"]
    clients = []
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        clients.append({
            "client_code": str(row[0]),
            "project_code": str(row[1]),
            "client_name": row[2],
            "project_name": row[3],
            "industry": row[4],
            "asset_size": row[5],
            "listing_status": row[6],
            "business_report": row[7],
            "gaap": row[8],
            "consolidated": row[9],
            "subsidiary_count": row[10],
            "internal_control": row[11],
            "initial_audit": row[12],
            "fiscal_start": row[13],
            "department": row[14],
            "el_name": row[15],
            "pm_name": row[16],
            "template_status": row[17],
            "el_empno": str(row[18] or ""),
            "pm_empno": str(row[19] or ""),
            "qrp_empno": str(row[20] or ""),
            "qrp_name": str(row[21] or ""),
            "group_code": str(row[22] or ""),
            "contract_hours": row[23] or 0,
            "qrp_hours": row[24] or 0,
            "rm_hours": row[25] or 0,
            "el_hours": row[26] or 0,
            "pm_hours": row[27] or 0,
            "ra_elpm_hours": row[28] or 0,
            "axdx_hours": row[29] or 0,
            "et_controllable_budget": row[30] or 0,
            "fulcrum_hours": row[31] or 0,
            "ra_staff_hours": row[32] or 0,
            "specialist_hours": row[33] or 0,
            "travel_hours": row[34] or 0,
            "total_budget_hours": row[38] or 0,
        })
    return clients


def _parse_project_db(wb) -> list[dict]:
    ws = wb["Project기본정보"]
    projects = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        projects.append({
            "project_code": str(row[0]),
            "client_code": str(row[0]).split("-")[0],
            "client_name": row[1],
            "project_name": row[2],
            "industry": row[3],
            "asset_size": row[4],
            "listing_status": row[5],
            "business_report": row[6],
            "gaap": row[7],
            "consolidated": row[8],
            "subsidiary_count": row[9],
            "internal_control": row[10],
            "initial_audit": row[11],
            "fiscal_start": row[12],
            "department": row[13],
            "el_name": row[14],
            "pm_name": row[15],
            "template_status": row[16],
            "el_empno": str(row[17] or ""),
            "pm_empno": str(row[18] or ""),
            "qrp_empno": str(row[19] or ""),
            "qrp_name": str(row[20] or ""),
            "group_code": str(row[21] or ""),
            "contract_hours": row[22] or 0,
            "qrp_hours": row[23] or 0,
            "rm_hours": row[24] or 0,
            "el_hours": row[25] or 0,
            "pm_hours": row[26] or 0,
            "ra_elpm_hours": row[27] or 0,
            "axdx_hours": row[28] or 0,
            "et_controllable_budget": row[29] or 0,
            "fulcrum_hours": row[30] or 0,
            "ra_staff_hours": row[31] or 0,
            "specialist_hours": row[32] or 0,
            "travel_hours": row[33] or 0,
            "total_budget_hours": row[37] or 0,
        })
    return projects


def _parse_budget_detail_db(wb) -> list[dict]:
    ws = wb["개인별Budget"]
    details = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        # emp_name 추출: "이지연(260657)" → "이지연", or raw name
        raw_name = row[5] or ""
        emp_name = str(raw_name).split("(")[0].strip() if raw_name else ""

        details.append({
            "project_code": str(row[0]),
            "project_name": row[1],
            "department": row[2],
            "budget_unit": row[3],
            "staff_department": row[4],
            "emp_name": emp_name,
            "empno": str(row[6]) if row[6] else "",
            "grade": row[7] or "",
            "year_month": row[8],
            "budget_hours": float(row[9]) if row[9] else 0,
        })
    return details
