"""Excel export 유틸리티 — PwC 스타일 적용."""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# PwC style constants
HEADER_FILL = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="맑은 고딕", size=10)
BODY_FONT_BOLD = Font(name="맑은 고딕", size=10, bold=True)
TOTAL_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
TOTAL_FONT = Font(name="맑은 고딕", size=10, bold=True)
RED_FONT = Font(name="맑은 고딕", size=10, color="D93954")
GREEN_FONT = Font(name="맑은 고딕", size=10, color="22992E")
ORANGE_FONT = Font(name="맑은 고딕", size=10, color="D04A02")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="E0E0E0"),
)
HEADER_BORDER = Border(
    bottom=Side(style="medium", color="D04A02"),
)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
NUMBER_FMT = "#,##0"
NUMBER_FMT_1 = "#,##0.0"
PCT_FMT = "0.0%"


def create_styled_excel(
    title: str,
    headers: list[str],
    rows: list[list],
    col_types: list[str] | None = None,
    total_row: list | None = None,
) -> io.BytesIO:
    """스타일 적용된 Excel 파일 생성.

    col_types: 각 컬럼 타입 ('str', 'num', 'pct', 'progress')
      - 'str': 텍스트 좌측 정렬
      - 'num': 숫자 우측 정렬, 천단위 콤마
      - 'num1': 소수점 1자리
      - 'pct': 퍼센트 (0.0%)
      - 'progress': 진행률 색상 (>110 빨강, >90 주황, 나머지 초록)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel 시트명 31자 제한

    num_cols = len(headers)
    if not col_types:
        col_types = ["str"] * num_cols

    # Header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = HEADER_BORDER

    # Data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data):
            if col_idx >= num_cols:
                break
            cell = ws.cell(row=row_idx, column=col_idx + 1, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            ct = col_types[col_idx] if col_idx < len(col_types) else "str"

            if ct in ("num", "num1"):
                cell.alignment = RIGHT_ALIGN
                cell.number_format = NUMBER_FMT if ct == "num" else NUMBER_FMT_1
            elif ct == "pct":
                cell.alignment = RIGHT_ALIGN
                if isinstance(value, (int, float)):
                    cell.value = value / 100  # openpyxl pct format expects 0~1
                    cell.number_format = PCT_FMT
            elif ct == "progress":
                cell.alignment = RIGHT_ALIGN
                if isinstance(value, (int, float)):
                    cell.number_format = "0.0\"%\""
                    if value > 110:
                        cell.font = Font(name="맑은 고딕", size=10, bold=True, color="D93954")
                    elif value > 90:
                        cell.font = Font(name="맑은 고딕", size=10, bold=True, color="D04A02")
                    else:
                        cell.font = Font(name="맑은 고딕", size=10, bold=True, color="22992E")
            else:
                cell.alignment = LEFT_ALIGN

    # Total row
    if total_row:
        total_row_idx = len(rows) + 2
        for col_idx, value in enumerate(total_row):
            if col_idx >= num_cols:
                break
            cell = ws.cell(row=total_row_idx, column=col_idx + 1, value=value)
            cell.font = TOTAL_FONT
            cell.fill = TOTAL_FILL
            cell.border = Border(top=Side(style="medium", color="2D2D2D"), bottom=Side(style="medium", color="2D2D2D"))
            ct = col_types[col_idx] if col_idx < len(col_types) else "str"
            if ct in ("num", "num1"):
                cell.alignment = RIGHT_ALIGN
                cell.number_format = NUMBER_FMT if ct == "num" else NUMBER_FMT_1
            elif ct == "pct":
                cell.alignment = RIGHT_ALIGN
                if isinstance(value, (int, float)):
                    cell.value = value / 100
                    cell.number_format = PCT_FMT
            elif ct == "progress":
                cell.alignment = RIGHT_ALIGN
                if isinstance(value, (int, float)):
                    cell.number_format = "0.0\"%\""
                    if value > 110:
                        cell.font = Font(name="맑은 고딕", size=10, bold=True, color="D93954")
                    elif value > 90:
                        cell.font = Font(name="맑은 고딕", size=10, bold=True, color="D04A02")
                    else:
                        cell.font = Font(name="맑은 고딕", size=10, bold=True, color="22992E")
            else:
                cell.alignment = LEFT_ALIGN

    # Auto column width
    for col_idx in range(1, num_cols + 1):
        max_len = len(str(headers[col_idx - 1])) * 1.5  # header tends to be shorter in Korean
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    cell_len = len(str(cell.value))
                    if cell_len > max_len:
                        max_len = cell_len
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto filter
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{ws.max_row}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
