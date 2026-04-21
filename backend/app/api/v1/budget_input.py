"""Budget 입력 API (3-Step Wizard)."""
import io

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional
from datetime import datetime

from app.db.session import get_db
from app.models.project import Client, Project, ServiceTaskMaster
from app.models.budget import BudgetDetail
from app.models.budget_master import (
    BudgetUnitMaster, PeerStatistics, PeerGroupMapping,
    ProjectMember, BudgetChangeLog,
)
from app.services.budget_service import upsert_project_from_client_data, bulk_insert_budget_details
from app.api.deps import get_optional_user, require_elpm, require_login, assert_can_modify_project, assert_can_delete_project

router = APIRouter()

# ── Service Types ────────────────────────────────────

SERVICE_TYPES = [
    {"code": "AUDIT", "name": "감사"},
    {"code": "AC", "name": "회계자문"},
    {"code": "IC", "name": "내부통제 (C.SOX PA)"},
    {"code": "ESG", "name": "ESG"},
    {"code": "VAL", "name": "Valuation"},
    {"code": "TRADE", "name": "통상자문"},
    {"code": "ACT", "name": "보험계리"},
    {"code": "ETC", "name": "기타"},
]


@router.get("/master/service-types")
def get_service_types():
    """서비스 분류 목록."""
    return SERVICE_TYPES


@router.get("/master/tasks")
def get_service_tasks(service_type: str = "AUDIT", db: Session = Depends(get_db)):
    """분류별 Task 마스터 목록."""
    rows = (
        db.query(ServiceTaskMaster)
        .filter(ServiceTaskMaster.service_type == service_type)
        .order_by(ServiceTaskMaster.sort_order)
        .all()
    )
    return [
        {
            "id": r.id,
            "service_type": r.service_type,
            "task_category": r.task_category or "",
            "task_name": r.task_name,
            "budget_unit_type": r.budget_unit_type or "",
            "sort_order": r.sort_order,
            "description": r.description or "",
            "activity_subcategory": r.activity_subcategory or "",
            "activity_detail": r.activity_detail or "",
            "budget_unit": r.budget_unit or "",
            "role": r.role or "",
        }
        for r in rows
    ]


@router.get("/master/activity-mapping")
def get_activity_mapping(service_type: str, db: Session = Depends(get_db)):
    """Step 2 구성원 Activity 매핑 드롭다운 소스."""
    rows = (
        db.query(ServiceTaskMaster)
        .filter(ServiceTaskMaster.service_type == service_type)
        .order_by(ServiceTaskMaster.sort_order)
        .all()
    )
    return [
        {
            "category": r.task_category or "",
            "subcategory": r.activity_subcategory or "",
            "detail": r.activity_detail or "",
            "role": r.role or "",
        }
        for r in rows
    ]


# ── Client Search ────────────────────────────────────

def _client_needs_detail(c) -> bool:
    """주요 상세필드가 모두 비어있으면 True — Azure sync 이후 사용자 입력 전 상태."""
    key_fields = [c.industry, c.asset_size, c.listing_status, c.gaap]
    return not any(f for f in key_fields)


@router.get("/clients/search")
def search_clients(q: str = "", db: Session = Depends(get_db)):
    """클라이언트 이름/코드로 검색."""
    query = db.query(Client)
    if q:
        query = query.filter(
            (Client.client_name.ilike(f"%{q}%")) |
            (Client.client_code.ilike(f"%{q}%"))
        )
    results = query.order_by(Client.client_name).limit(50).all()
    return [
        {
            "client_code": c.client_code,
            "client_name": c.client_name or "",
            "industry": c.industry or "",
            "asset_size": c.asset_size or "",
            "listing_status": c.listing_status or "",
            "gaap": c.gaap or "",
            "consolidated": c.consolidated or "",
            "subsidiary_count": c.subsidiary_count or "",
            "internal_control": c.internal_control or "",
            "business_report": c.business_report or "",
            "initial_audit": c.initial_audit or "",
            "needs_detail": _client_needs_detail(c),
        }
        for c in results
    ]


@router.get("/clients/{client_code}/info")
def get_client_info(
    client_code: str,
    user: dict = Depends(require_login),
    db: Session = Depends(get_db),
):
    """클라이언트 기본 정보 단건 조회 (Step 1 자동입력용)."""
    c = db.query(Client).filter(Client.client_code == client_code).first()
    if c is None:
        raise HTTPException(status_code=404, detail="클라이언트를 찾을 수 없습니다.")
    return {
        "client_code": c.client_code,
        "client_name": c.client_name,
        "industry": c.industry or "",
        "asset_size": c.asset_size or "",
        "listing_status": c.listing_status or "",
        "business_report": c.business_report or "",
        "gaap": c.gaap or "",
        "consolidated": c.consolidated or "",
        "subsidiary_count": c.subsidiary_count or "",
        "internal_control": c.internal_control or "",
        "initial_audit": c.initial_audit or "",
    }


@router.get("/employees/search")
def search_employees(
    q: str = "",
    include_inactive: bool = False,
    db: Session = Depends(get_db),
):
    """직원 이름/사번으로 검색 — Postgres employees 마스터 단일 조회.

    기본값으로 재직 직원만 반환. include_inactive=true 시 퇴사/휴직 포함.
    """
    if not q or len(q) < 2:
        return []
    from app.models.employee import Employee
    query = db.query(Employee).filter(
        (Employee.name.ilike(f"%{q}%")) |
        (Employee.empno.ilike(f"%{q}%"))
    )
    if not include_inactive:
        query = query.filter(Employee.emp_status == "재직")
    rows = query.order_by(Employee.name).limit(30).all()
    return [
        {
            "empno": e.empno,
            "name": e.name,
            "grade": e.grade_name or "",
            "department": e.department or "",
            "emp_status": e.emp_status or "",
        }
        for e in rows
    ]


@router.get("/projects/list")
def list_registered_projects(
    q: str = "",
    db: Session = Depends(get_db),
    user: Optional[dict] = Depends(get_optional_user),
):
    """Budget 등록된 프로젝트 목록 — 로그인 사용자가 EL인 프로젝트만."""
    from sqlalchemy import func as sa_func
    query = db.query(Project)
    if user:
        query = query.filter(Project.el_empno == user["empno"])
    if q:
        query = query.filter(
            (Project.project_name.ilike(f"%{q}%")) |
            (Project.project_code.ilike(f"%{q}%"))
        )
    projects = query.order_by(Project.contract_hours.desc().nullslast()).all()
    result = []
    for p in projects:
        member_count = db.query(sa_func.count(ProjectMember.id)).filter(
            ProjectMember.project_code == p.project_code).scalar() or 0
        result.append({
            "project_code": p.project_code,
            "project_name": p.project_name or "",
            "el_name": p.el_name or "",
            "pm_name": p.pm_name or "",
            "contract_hours": float(p.contract_hours or 0),
            "total_budget_hours": float(p.total_budget_hours or 0),
            "template_status": p.template_status or "작성중",
            "member_count": member_count,
        })
    return result


@router.get("/projects/search")
def search_projects(q: str = "", client_code: str = "", db: Session = Depends(get_db)):
    """프로젝트 검색 — Azure DB(회사 전체) + PostgreSQL(Budget 등록 여부) 병합.
    client_code가 주어지면 해당 클라이언트(프로젝트코드 앞 5자리)에 종속된 프로젝트만 반환.
    """
    from app.services import azure_service

    # 1) Azure에서 회사 전체 진행 중 프로젝트 검색
    #    client_code가 있으면 해당 클라이언트의 프로젝트만 필터
    azure_results = azure_service.search_azure_projects(
        q, limit=200, client_code_prefix=client_code
    )

    # 2) PostgreSQL에서 이미 Budget 등록된 프로젝트 조회
    registered_codes = set()
    pg_data: dict[str, dict] = {}
    if azure_results:
        codes = [r["project_code"] for r in azure_results]
        pg_projects = db.query(Project).filter(Project.project_code.in_(codes)).all()
        for p in pg_projects:
            registered_codes.add(p.project_code)
            pg_data[p.project_code] = {
                "contract_hours": float(p.contract_hours or 0),
                "axdx_hours": float(p.axdx_hours or 0),
                "qrp_hours": float(p.qrp_hours or 0),
                "rm_hours": float(p.rm_hours or 0),
                "el_hours": float(p.el_hours or 0),
                "pm_hours": float(p.pm_hours or 0),
                "ra_elpm_hours": float(p.ra_elpm_hours or 0),
                "et_controllable_budget": float(p.et_controllable_budget or 0),
                "fulcrum_hours": float(p.fulcrum_hours or 0),
                "ra_staff_hours": float(p.ra_staff_hours or 0),
                "specialist_hours": float(p.specialist_hours or 0),
                "travel_hours": float(p.travel_hours or 0),
                "total_budget_hours": float(p.total_budget_hours or 0),
                "template_status": p.template_status or "작성중",
                "qrp_name": p.qrp_name or "",
                "qrp_empno": p.qrp_empno or "",
            }

    # 3) 병합: Azure 기본정보 + PG Budget 정보
    return [
        {
            "project_code": r["project_code"],
            "project_name": r["project_name"],
            "client_name": r["client_name"],
            "department": r["department"],
            "el_name": r["el_name"],
            "el_empno": r["el_empno"],
            "pm_name": r["pm_name"],
            "pm_empno": r["pm_empno"],
            "industry": r["industry"],
            "is_registered": r["project_code"] in registered_codes,
            **pg_data.get(r["project_code"], {
                "contract_hours": 0, "axdx_hours": 0, "qrp_hours": 0,
                "rm_hours": 0, "el_hours": 0, "pm_hours": 0,
                "ra_elpm_hours": 0, "et_controllable_budget": 0,
                "fulcrum_hours": 0, "ra_staff_hours": 0,
                "specialist_hours": 0, "travel_hours": 0,
                "total_budget_hours": 0, "template_status": "작성중",
                "qrp_name": "", "qrp_empno": "",
            }),
        }
        for r in azure_results
    ]


@router.get("/projects/{project_code}/clone-data")
def get_clone_data(project_code: str, db: Session = Depends(get_db)):
    """이전 프로젝트 정보 가져오기 — 시간, 구성원, budget template 전체 반환."""
    from collections import defaultdict

    proj = db.query(Project).filter(Project.project_code == project_code).first()
    if not proj:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다")

    # 시간 정보
    hours = {
        "contract_hours": float(proj.contract_hours or 0),
        "axdx_hours": float(proj.axdx_hours or 0),
        "qrp_hours": float(proj.qrp_hours or 0),
        "rm_hours": float(proj.rm_hours or 0),
        "el_hours": float(proj.el_hours or 0),
        "pm_hours": float(proj.pm_hours or 0),
        "ra_elpm_hours": float(proj.ra_elpm_hours or 0),
        "et_controllable_budget": float(proj.et_controllable_budget or 0),
        "fulcrum_hours": float(proj.fulcrum_hours or 0),
        "ra_staff_hours": float(proj.ra_staff_hours or 0),
        "specialist_hours": float(proj.specialist_hours or 0),
        "travel_hours": float(proj.travel_hours or 0),
        "total_budget_hours": float(proj.total_budget_hours or 0),
    }

    # 구성원
    members = (
        db.query(ProjectMember)
        .filter(ProjectMember.project_code == project_code)
        .order_by(ProjectMember.sort_order)
        .all()
    )
    members_data = [
        {"role": m.role, "name": m.name, "empno": m.empno, "grade": m.grade or "",
         "activity_mapping": m.activity_mapping, "sort_order": m.sort_order}
        for m in members
    ]

    # Budget template
    details = (
        db.query(BudgetDetail)
        .filter(BudgetDetail.project_code == project_code)
        .all()
    )
    rows_map = defaultdict(lambda: {"months": {}, "meta": {}})
    for d in details:
        key = (d.budget_category, d.budget_unit, d.empno)
        rows_map[key]["meta"] = {
            "budget_category": d.budget_category,
            "budget_unit": d.budget_unit,
            "empno": d.empno,
            "emp_name": d.emp_name,
            "grade": d.grade,
            "department": d.department,
        }
        if d.budget_hours and d.budget_hours > 0:
            rows_map[key]["months"][d.year_month] = d.budget_hours
    template_rows = []
    for key, val in rows_map.items():
        total = sum(val["months"].values())
        template_rows.append({**val["meta"], "months": val["months"], "total": total})

    return {
        "project_code": project_code,
        "project_name": proj.project_name,
        "hours": hours,
        "members": members_data,
        "template": {"rows": template_rows},
    }


# ── Schemas ──────────────────────────────────────────

class ProjectCreateRequest(BaseModel):
    project_code: str
    project_name: str
    client_code: Optional[str] = None
    client_name: Optional[str] = None
    department: Optional[str] = ""
    industry: Optional[str] = None
    asset_size: Optional[str] = None
    listing_status: Optional[str] = None
    business_report: Optional[str] = None
    gaap: Optional[str] = None
    consolidated: Optional[str] = None
    subsidiary_count: Optional[str] = None
    internal_control: Optional[str] = None
    initial_audit: Optional[str] = None
    group_code: Optional[str] = None
    el_empno: Optional[str] = ""
    el_name: Optional[str] = ""
    pm_empno: Optional[str] = ""
    pm_name: Optional[str] = ""
    qrp_empno: Optional[str] = ""
    qrp_name: Optional[str] = ""
    contract_hours: float = 0
    axdx_hours: float = 0
    qrp_hours: float = 0
    rm_hours: float = 0
    el_hours: float = 0
    pm_hours: float = 0
    ra_elpm_hours: float = 0
    et_controllable_budget: float = 0
    fulcrum_hours: float = 0
    ra_staff_hours: float = 0
    specialist_hours: float = 0
    travel_hours: float = 0
    total_budget_hours: float = 0
    template_status: Optional[str] = "작성중"
    fiscal_start: Optional[str] = None  # "2025-04"
    service_type: Optional[str] = "AUDIT"


class MemberRequest(BaseModel):
    role: str  # "FLDT 구성원" | "지원 ET 구성원"
    name: str
    empno: Optional[str] = ""
    grade: Optional[str] = ""
    activity_mapping: Optional[str] = "재무제표기말감사"
    sort_order: int = 0


class BudgetTemplateRow(BaseModel):
    budget_category: str
    budget_unit: str
    empno: Optional[str] = ""
    emp_name: Optional[str] = ""
    grade: Optional[str] = ""
    months: dict[str, float] = {}  # {"2025-06": 8, "2025-12": 4, ...}


class BudgetTemplateSaveRequest(BaseModel):
    rows: list[BudgetTemplateRow]
    template_status: str = "작성중"


# ── Step 1: 프로젝트 기본정보 ────────────────────────

@router.post("/projects")
def create_project(
    req: ProjectCreateRequest,
    user: dict = Depends(require_elpm),
    db: Session = Depends(get_db),
):
    """프로젝트 생성 (Step 1) — EL/PM 또는 관리자만 가능."""
    # 관리자가 아닌 경우, 본인이 EL 또는 PM 인 프로젝트만 생성 가능
    if user["role"] != "admin":
        if user["empno"] not in (req.el_empno or "", req.pm_empno or ""):
            raise HTTPException(
                status_code=403,
                detail="본인이 EL 또는 PM 인 프로젝트만 생성할 수 있습니다.",
            )

    existing = db.query(Project).filter(Project.project_code == req.project_code).first()
    if existing:
        raise HTTPException(400, f"Project '{req.project_code}' already exists. Use PUT to update.")

    data = req.model_dump()
    data["client_code"] = data.get("client_code") or data["project_code"].split("-")[0]
    project = upsert_project_from_client_data(db, data)

    BudgetChangeLog(
        project_code=req.project_code,
        change_type="create",
        change_summary=f"프로젝트 생성: {req.project_name}",
    )
    db.add(BudgetChangeLog(
        project_code=req.project_code,
        change_type="create",
        change_summary=f"프로젝트 생성: {req.project_name}",
    ))
    db.commit()

    return {"message": "프로젝트 생성 완료", "project_code": project.project_code}


@router.put("/projects/{project_code}")
def update_project(
    project_code: str,
    req: ProjectCreateRequest,
    user: dict = Depends(require_elpm),
    db: Session = Depends(get_db),
):
    """프로젝트 수정 (Step 1) — EL/PM 또는 관리자만 가능."""
    assert_can_modify_project(db, user, project_code)
    data = req.model_dump()
    data["project_code"] = project_code
    data["client_code"] = data.get("client_code") or project_code.split("-")[0]
    project = upsert_project_from_client_data(db, data)

    db.add(BudgetChangeLog(
        project_code=project_code,
        change_type="update",
        change_summary="기본정보 수정",
    ))
    db.commit()

    return {"message": "프로젝트 수정 완료", "project_code": project.project_code}


@router.delete("/projects/{project_code}")
def delete_project(
    project_code: str,
    user: dict = Depends(require_elpm),
    db: Session = Depends(get_db),
):
    """프로젝트 삭제 (EL 또는 관리자만 가능)."""
    assert_can_delete_project(db, user, project_code)

    proj = db.query(Project).filter(Project.project_code == project_code).first()
    if not proj:
        raise HTTPException(status_code=404, detail="프로젝트를 찾을 수 없습니다")

    # 관련 데이터 삭제
    db.query(BudgetDetail).filter(BudgetDetail.project_code == project_code).delete()
    db.query(ProjectMember).filter(ProjectMember.project_code == project_code).delete()
    db.query(BudgetChangeLog).filter(BudgetChangeLog.project_code == project_code).delete()

    # Client 삭제 (다른 프로젝트에서 참조하지 않는 경우)
    if proj.client_id:
        other = db.query(Project).filter(
            Project.client_id == proj.client_id, Project.project_code != project_code
        ).count()
        if other == 0:
            db.query(Client).filter(Client.id == proj.client_id).delete()

    db.delete(proj)
    db.commit()

    return {"message": f"{project_code} 삭제 완료"}


@router.get("/projects/{project_code}/info")
def get_project_info(project_code: str, db: Session = Depends(get_db)):
    """프로젝트 기본정보 조회."""
    project = db.query(Project).filter(Project.project_code == project_code).first()
    if not project:
        raise HTTPException(404, "Project not found")

    client = db.query(Client).filter(Client.id == project.client_id).first()

    return {
        "project": {
            "project_code": project.project_code,
            "project_name": project.project_name,
            "department": project.department,
            "el_empno": project.el_empno,
            "el_name": project.el_name,
            "pm_empno": project.pm_empno,
            "pm_name": project.pm_name,
            "contract_hours": project.contract_hours,
            "axdx_hours": project.axdx_hours,
            "qrp_hours": project.qrp_hours,
            "rm_hours": project.rm_hours,
            "el_hours": project.el_hours,
            "pm_hours": project.pm_hours,
            "ra_elpm_hours": project.ra_elpm_hours,
            "et_controllable_budget": project.et_controllable_budget,
            "fulcrum_hours": project.fulcrum_hours,
            "ra_staff_hours": project.ra_staff_hours,
            "specialist_hours": project.specialist_hours,
            "travel_hours": project.travel_hours,
            "total_budget_hours": project.total_budget_hours,
            "template_status": project.template_status,
            "service_type": project.service_type or "AUDIT",
        },
        "client": {
            "client_code": client.client_code if client else "",
            "client_name": client.client_name if client else "",
            "industry": client.industry if client else "",
            "asset_size": client.asset_size if client else "",
            "listing_status": client.listing_status if client else "",
            "business_report": client.business_report if client else "",
            "gaap": client.gaap if client else "",
            "consolidated": client.consolidated if client else "",
            "subsidiary_count": client.subsidiary_count if client else "",
            "internal_control": client.internal_control if client else "",
            "initial_audit": client.initial_audit if client else "",
        } if client else None,
    }


# ── Step 2: 구성원 관리 ─────────────────────────────

# NOTE: export/upload routes must be registered BEFORE the plain
# /members GET so that FastAPI does not swallow the sub-path as a
# path-parameter match against {project_code}/members.

@router.get("/projects/{project_code}/members/export")
def export_project_members(
    project_code: str,
    user: dict = Depends(require_login),
    db: Session = Depends(get_db),
):
    """Step 2 구성원 목록 Excel 다운로드.

    NOTE: This route MUST be registered before the plain `/members` GET so
    that FastAPI does not swallow `members/export` as a {project_code}
    sub-path match.
    """
    from openpyxl import Workbook

    members = (
        db.query(ProjectMember)
        .filter(ProjectMember.project_code == project_code)
        .order_by(ProjectMember.sort_order, ProjectMember.id)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "구성원"
    ws.append(["empno", "name", "role", "grade"])
    for m in members:
        ws.append([m.empno or "", m.name or "", m.role or "FLDT", m.grade or ""])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="members_{project_code}.xlsx"',
        },
    )


@router.get("/projects/{project_code}/members")
def get_members(project_code: str, db: Session = Depends(get_db)):
    members = (
        db.query(ProjectMember)
        .filter(ProjectMember.project_code == project_code)
        .order_by(ProjectMember.sort_order)
        .all()
    )
    return [
        {
            "id": m.id,
            "role": m.role,
            "name": m.name,
            "empno": m.empno,
            "grade": m.grade or "",
            "activity_mapping": m.activity_mapping,
            "sort_order": m.sort_order,
        }
        for m in members
    ]


@router.put("/projects/{project_code}/members")
def save_members(
    project_code: str,
    members: list[MemberRequest],
    user: dict = Depends(require_elpm),
    db: Session = Depends(get_db),
):
    """구성원 일괄 저장 (기존 삭제 후 재삽입) — EL/PM 또는 관리자만 가능."""
    assert_can_modify_project(db, user, project_code)
    db.query(ProjectMember).filter(ProjectMember.project_code == project_code).delete()
    for i, m in enumerate(members):
        db.add(ProjectMember(
            project_code=project_code,
            role=m.role,
            name=m.name,
            empno=m.empno,
            grade=m.grade or "",
            activity_mapping=m.activity_mapping,
            sort_order=m.sort_order or i,
        ))
    db.commit()
    return {"message": f"{len(members)}명 구성원 저장 완료"}


# ── Step 3: Budget Template ──────────────────────────

@router.get("/projects/{project_code}/template")
def get_template(project_code: str, db: Session = Depends(get_db)):
    """Budget Template 조회."""
    details = (
        db.query(BudgetDetail)
        .filter(BudgetDetail.project_code == project_code)
        .all()
    )

    # 행 그룹핑: (category, unit, empno) → months
    from collections import defaultdict
    rows_map = defaultdict(lambda: {"months": {}, "meta": {}})
    for d in details:
        key = (d.budget_category, d.budget_unit, d.empno)
        rows_map[key]["meta"] = {
            "budget_category": d.budget_category,
            "budget_unit": d.budget_unit,
            "empno": d.empno,
            "emp_name": d.emp_name,
            "grade": d.grade,
            "department": d.department,
        }
        if d.budget_hours and d.budget_hours > 0:
            rows_map[key]["months"][d.year_month] = d.budget_hours

    rows = []
    for key, val in rows_map.items():
        total = sum(val["months"].values())
        rows.append({
            **val["meta"],
            "months": val["months"],
            "total": total,
        })

    return {"project_code": project_code, "rows": rows}


@router.put("/projects/{project_code}/template")
def save_template(
    project_code: str,
    req: BudgetTemplateSaveRequest,
    user: dict = Depends(require_elpm),
    db: Session = Depends(get_db),
):
    """Budget Template 저장 — EL/PM 또는 관리자만 가능."""
    assert_can_modify_project(db, user, project_code)
    # budget_details로 변환
    details = []
    for row in req.rows:
        for ym, hours in row.months.items():
            if hours and hours > 0:
                details.append({
                    "budget_category": row.budget_category,
                    "budget_unit": row.budget_unit,
                    "empno": row.empno,
                    "emp_name": row.emp_name,
                    "grade": row.grade,
                    "year_month": ym,
                    "budget_hours": hours,
                })

    bulk_insert_budget_details(db, project_code, details)

    # template_status 업데이트
    project = db.query(Project).filter(Project.project_code == project_code).first()
    if project:
        project.template_status = req.template_status

    db.add(BudgetChangeLog(
        project_code=project_code,
        change_type="update",
        change_summary=f"Budget Template 저장 ({len(details)}건, 상태: {req.template_status})",
    ))
    db.commit()

    return {"message": f"Budget Template 저장 완료 ({len(details)}건)"}


# ── 마스터 데이터 ────────────────────────────────────

@router.get("/master/units")
def get_budget_units(db: Session = Depends(get_db)):
    """관리단위 마스터 목록."""
    units = db.query(BudgetUnitMaster).order_by(BudgetUnitMaster.sort_order).all()
    if not units:
        # DB에 없으면 하드코딩 기본값 반환
        return {"units": DEFAULT_BUDGET_UNITS}
    return {
        "units": [
            {"category": u.category, "unit_name": u.unit_name, "sort_order": u.sort_order}
            for u in units
        ]
    }


@router.get("/peer-group")
def get_peer_group(
    industry: str = "",
    asset_size: str = "",
    listing_status: str = "",
    consolidated: str = "",
    internal_control: str = "",
    db: Session = Depends(get_db),
):
    """유사회사 그룹 조회."""
    mapping = (
        db.query(PeerGroupMapping)
        .filter(
            PeerGroupMapping.industry == industry,
            PeerGroupMapping.asset_size == asset_size,
            PeerGroupMapping.listing_status == listing_status,
            PeerGroupMapping.consolidated == consolidated,
            PeerGroupMapping.internal_control == internal_control,
        )
        .first()
    )
    if not mapping:
        return {"stat_group": None}
    return {"stat_group": mapping.stat_group}


@router.get("/master/peer-stats")
def get_peer_stats(group: str, db: Session = Depends(get_db)):
    """유사회사 통계."""
    stats = (
        db.query(PeerStatistics)
        .filter(PeerStatistics.stat_group == group)
        .all()
    )
    return {
        "stats": {s.budget_unit: s.avg_ratio for s in stats}
    }


@router.get("/projects/{project_code}/history")
def get_change_history(project_code: str, db: Session = Depends(get_db)):
    """변경 이력."""
    logs = (
        db.query(BudgetChangeLog)
        .filter(BudgetChangeLog.project_code == project_code)
        .order_by(BudgetChangeLog.changed_at.desc())
        .limit(50)
        .all()
    )
    return [
        {
            "changed_at": l.changed_at.isoformat() if l.changed_at else "",
            "changed_by": l.changed_by_name or l.changed_by_empno or "",
            "change_type": l.change_type,
            "change_summary": l.change_summary,
        }
        for l in logs
    ]


# ── 기본 관리단위 목록 (DB 비어있을 때 사용) ─────────

DEFAULT_BUDGET_UNITS = [
    {"category": "분반기 검토", "unit_name": "분반기 검토", "sort_order": 1},
    {"category": "계획단계", "unit_name": "계획단계", "sort_order": 10},
    {"category": "계획단계", "unit_name": "초도감사", "sort_order": 11},
    {"category": "재무제표 수준 위험", "unit_name": "부정위험", "sort_order": 20},
    {"category": "재무제표 수준 위험", "unit_name": "계속기업", "sort_order": 21},
    {"category": "재무제표 수준 위험", "unit_name": "기타(특수관계자/법률 및 규정)", "sort_order": 22},
    {"category": "자산", "unit_name": "현금및현금성자산-일반", "sort_order": 30},
    {"category": "자산", "unit_name": "현금및현금성자산-조회", "sort_order": 31},
    {"category": "자산", "unit_name": "채무 및 지분증권-일반", "sort_order": 32},
    {"category": "자산", "unit_name": "채무 및 지분증권-공정가치평가", "sort_order": 33},
    {"category": "자산", "unit_name": "파생상품", "sort_order": 34},
    {"category": "자산", "unit_name": "매출채권-일반", "sort_order": 35},
    {"category": "자산", "unit_name": "매출채권-조회", "sort_order": 36},
    {"category": "자산", "unit_name": "고객 반품, 할인 및 충당금", "sort_order": 37},
    {"category": "자산", "unit_name": "재고자산-일반", "sort_order": 38},
    {"category": "자산", "unit_name": "재고자산-실사입회및문서화", "sort_order": 39},
    {"category": "자산", "unit_name": "건설계약", "sort_order": 40},
    {"category": "자산", "unit_name": "기타자산", "sort_order": 41},
    {"category": "자산", "unit_name": "유/무형자산/투자부동산-일반", "sort_order": 42},
    {"category": "자산", "unit_name": "유/무형자산/투자부동산-취득처분", "sort_order": 43},
    {"category": "자산", "unit_name": "종속기업 등에 대한 투자자산-일반", "sort_order": 44},
    {"category": "자산", "unit_name": "종속기업 등에 대한 투자자산-손상검토", "sort_order": 45},
    {"category": "자산", "unit_name": "영업권-일반", "sort_order": 46},
    {"category": "자산", "unit_name": "영업권-PPA", "sort_order": 47},
    {"category": "자산", "unit_name": "영업권-손상검토", "sort_order": 48},
    {"category": "자산", "unit_name": "자산 - 기타", "sort_order": 49},
    {"category": "부채 및 자본", "unit_name": "매입채무-일반", "sort_order": 50},
    {"category": "부채 및 자본", "unit_name": "매입채무-조회", "sort_order": 51},
    {"category": "부채 및 자본", "unit_name": "특수관계자/내부거래", "sort_order": 52},
    {"category": "부채 및 자본", "unit_name": "기타부채 및 충당부채-일반", "sort_order": 53},
    {"category": "부채 및 자본", "unit_name": "기타부채 및 충당부채-부외부채/우발부채검토", "sort_order": 54},
    {"category": "부채 및 자본", "unit_name": "법인세", "sort_order": 55},
    {"category": "부채 및 자본", "unit_name": "차입금/복합금융상품-일반", "sort_order": 56},
    {"category": "부채 및 자본", "unit_name": "차입금/복합금융상품-공정가치평가", "sort_order": 57},
    {"category": "부채 및 자본", "unit_name": "차입금/복합금융상품-약정사항 검토", "sort_order": 58},
    {"category": "부채 및 자본", "unit_name": "이연수익", "sort_order": 59},
    {"category": "부채 및 자본", "unit_name": "리스", "sort_order": 60},
    {"category": "부채 및 자본", "unit_name": "퇴직급여/장기종업원급여", "sort_order": 61},
    {"category": "부채 및 자본", "unit_name": "부채 - 기타", "sort_order": 62},
    {"category": "부채 및 자본", "unit_name": "자본금 및 기타 자본 계정", "sort_order": 63},
    {"category": "수익/비용", "unit_name": "매출(IIFRS15/K-GAAP)-일반", "sort_order": 70},
    {"category": "수익/비용", "unit_name": "매출(IIFRS15/K-GAAP)-발생사실테스트", "sort_order": 71},
    {"category": "수익/비용", "unit_name": "매출(IIFRS15/K-GAAP)-Cut-off테스트", "sort_order": 72},
    {"category": "수익/비용", "unit_name": "기타 영업수익", "sort_order": 73},
    {"category": "수익/비용", "unit_name": "매출원가-일반", "sort_order": 74},
    {"category": "수익/비용", "unit_name": "매출원가-증빙테스트", "sort_order": 75},
    {"category": "수익/비용", "unit_name": "영업비용", "sort_order": 76},
    {"category": "수익/비용", "unit_name": "인건비", "sort_order": 77},
    {"category": "수익/비용", "unit_name": "영업외손익", "sort_order": 78},
    {"category": "수익/비용", "unit_name": "주식기준보상비용-일반", "sort_order": 79},
    {"category": "수익/비용", "unit_name": "비용 - 기타", "sort_order": 80},
    {"category": "종결단계", "unit_name": "종결단계", "sort_order": 90},
    {"category": "종결단계", "unit_name": "주석검토-별도", "sort_order": 91},
    {"category": "종결단계", "unit_name": "주석검토-연결", "sort_order": 92},
    {"category": "종결단계", "unit_name": "기말감사 - 별도CF", "sort_order": 93},
    {"category": "종결단계", "unit_name": "외국어 보고서", "sort_order": 94},
    {"category": "연결", "unit_name": "기말감사 - 연결일반", "sort_order": 100},
    {"category": "연결", "unit_name": "기말감사 - 연결GA/CA", "sort_order": 101},
    {"category": "연결", "unit_name": "기말감사 - 연결법인세", "sort_order": 102},
    {"category": "연결", "unit_name": "기말감사 - 연결CF", "sort_order": 103},
    {"category": "내부통제", "unit_name": "내부통제-계획/종결", "sort_order": 110},
    {"category": "내부통제", "unit_name": "내부회계검토", "sort_order": 111},
    {"category": "내부통제", "unit_name": "설계평가-내부통제-ELC", "sort_order": 112},
    {"category": "내부통제", "unit_name": "설계평가-내부통제-ITGC", "sort_order": 113},
    {"category": "내부통제", "unit_name": "설계평가-내부통제-재무보고(FR)", "sort_order": 114},
    {"category": "내부통제", "unit_name": "설계평가-내부통제-자금", "sort_order": 115},
    {"category": "내부통제", "unit_name": "설계평가-내부통제-매출/매출채권", "sort_order": 116},
    {"category": "내부통제", "unit_name": "설계평가-내부통제-재고/원가", "sort_order": 117},
    {"category": "내부통제", "unit_name": "설계평가-내부통제-매입/매입채무", "sort_order": 118},
    {"category": "내부통제", "unit_name": "설계평가-내부통제-급여/퇴직급여", "sort_order": 119},
    {"category": "내부통제", "unit_name": "설계평가-내부통제-유/무형자산/투자부동산", "sort_order": 120},
    {"category": "내부통제", "unit_name": "설계평가-내부통제-연결", "sort_order": 121},
    {"category": "내부통제", "unit_name": "설계평가-내부통제-기타프로세스", "sort_order": 122},
    {"category": "내부통제", "unit_name": "운영평가-내부통제-ELC", "sort_order": 130},
    {"category": "내부통제", "unit_name": "운영평가-내부통제-ITGC", "sort_order": 131},
    {"category": "내부통제", "unit_name": "운영평가-내부통제-재무보고(FR)", "sort_order": 132},
    {"category": "내부통제", "unit_name": "운영평가-내부통제-자금", "sort_order": 133},
    {"category": "내부통제", "unit_name": "운영평가-내부통제-매출/매출채권", "sort_order": 134},
    {"category": "내부통제", "unit_name": "운영평가-내부통제-재고/원가", "sort_order": 135},
    {"category": "내부통제", "unit_name": "운영평가-내부통제-매입/매입채무", "sort_order": 136},
    {"category": "내부통제", "unit_name": "운영평가-내부통제-급여/퇴직급여", "sort_order": 137},
    {"category": "내부통제", "unit_name": "운영평가-내부통제-유/무형자산/투자부동산", "sort_order": 138},
    {"category": "내부통제", "unit_name": "운영평가-내부통제-연결", "sort_order": 139},
    {"category": "내부통제", "unit_name": "운영평가-내부통제-기타프로세스", "sort_order": 140},
    {"category": "IT 감사-RA", "unit_name": "IT 감사-RA", "sort_order": 150},
]


# ── Step 2 구성원 Excel Upload ──────────────────


@router.post("/projects/{project_code}/members/upload")
async def upload_project_members(
    project_code: str,
    file: UploadFile = File(...),
    user: dict = Depends(require_elpm),
    db: Session = Depends(get_db),
):
    """Step 2 구성원 목록 Excel 업로드 — FLDT 구성원 truncate+insert."""
    from openpyxl import load_workbook
    from app.models.employee import Employee

    assert_can_modify_project(db, user, project_code)

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))

    db.query(ProjectMember).filter(
        ProjectMember.project_code == project_code,
        ProjectMember.role == "FLDT",
    ).delete(synchronize_session=False)
    db.commit()

    imported: list[dict] = []
    skipped: list[dict] = []
    for idx, row in enumerate(rows):
        if not row or not any(c is not None for c in row):
            continue
        empno = str(row[0]).strip() if row[0] else ""
        name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        role = str(row[2]).strip() if len(row) > 2 and row[2] else "FLDT"
        grade = str(row[3]).strip() if len(row) > 3 and row[3] else ""

        if not empno:
            skipped.append({"row": idx + 2, "reason": "empno missing"})
            continue

        emp = db.query(Employee).filter(Employee.empno == empno).first()
        if emp is None:
            skipped.append({"empno": empno, "reason": "not_found"})
            continue
        if emp.emp_status and emp.emp_status != "재직":
            skipped.append({"empno": empno, "reason": "inactive"})
            continue

        db.add(ProjectMember(
            project_code=project_code,
            role=role,
            name=name or emp.name,
            empno=empno,
            grade=grade or (emp.grade_name or ""),
            sort_order=idx,
        ))
        imported.append({"empno": empno, "name": name or emp.name})

    db.commit()
    return {
        "imported_count": len(imported),
        "imported": imported,
        "skipped": skipped,
    }
