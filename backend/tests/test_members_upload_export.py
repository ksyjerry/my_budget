"""Tests for Step 2 members Excel export/upload (#40)."""
import io

import pytest
from openpyxl import Workbook, load_workbook

from app.db.session import SessionLocal
from app.models.budget_master import ProjectMember
from app.models.project import Project
from app.models.employee import Employee


TEST_PROJECT_CODE = "S4_TEST_MEM"


@pytest.fixture(autouse=True)
def _seed():
    s = SessionLocal()
    try:
        if s.query(Project).filter(Project.project_code == TEST_PROJECT_CODE).first() is None:
            s.add(Project(
                project_code=TEST_PROJECT_CODE,
                project_name="S4 Members Test",
                el_empno="170661", pm_empno="170661",
                contract_hours=100,
            ))
        for empno, name, status in [
            ("S4MB1", "멤버1", "재직"),
            ("S4MB2", "멤버2", "재직"),
            ("S4MB3", "퇴사자", "퇴사"),
        ]:
            if s.query(Employee).filter(Employee.empno == empno).first() is None:
                s.add(Employee(empno=empno, name=name, emp_status=status))
        s.query(ProjectMember).filter(ProjectMember.project_code == TEST_PROJECT_CODE).delete()
        s.add(ProjectMember(project_code=TEST_PROJECT_CODE, empno="S4MB1", name="멤버1", role="FLDT", grade="SA"))
        s.commit()
    finally:
        s.close()
    yield


def test_export_returns_xlsx_content_type(client, elpm_cookie):
    r = client.get(
        f"/api/v1/budget/projects/{TEST_PROJECT_CODE}/members/export",
        cookies=elpm_cookie,
    )
    assert r.status_code == 200
    assert "spreadsheetml.sheet" in r.headers.get("content-type", "")
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["사번", "이름", "역할", "직급", "팀"]


def test_upload_replaces_fldt_members(client, elpm_cookie):
    wb = Workbook()
    ws = wb.active
    ws.append(["empno", "name", "role", "grade"])
    ws.append(["S4MB2", "멤버2", "FLDT", "M"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post(
        f"/api/v1/budget/projects/{TEST_PROJECT_CODE}/members/upload",
        files={"file": ("members.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        cookies=elpm_cookie,
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["imported_count"] == 1
    s = SessionLocal()
    try:
        remaining = s.query(ProjectMember).filter(
            ProjectMember.project_code == TEST_PROJECT_CODE,
            ProjectMember.role == "FLDT",
        ).all()
        assert len(remaining) == 1
        assert remaining[0].empno == "S4MB2"
    finally:
        s.close()


def test_upload_skips_inactive_empno(client, elpm_cookie):
    wb = Workbook()
    ws = wb.active
    ws.append(["empno", "name", "role", "grade"])
    ws.append(["S4MB3", "퇴사자", "FLDT", "SA"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post(
        f"/api/v1/budget/projects/{TEST_PROJECT_CODE}/members/upload",
        files={"file": ("members.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        cookies=elpm_cookie,
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["imported_count"] == 0
    assert any(s.get("reason") == "inactive" for s in body["skipped"])


def test_staff_cannot_upload_members(client, staff_cookie):
    wb = Workbook()
    ws = wb.active
    ws.append(["empno", "name", "role", "grade"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post(
        f"/api/v1/budget/projects/{TEST_PROJECT_CODE}/members/upload",
        files={"file": ("members.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        cookies=staff_cookie,
    )
    assert r.status_code in (401, 403)
