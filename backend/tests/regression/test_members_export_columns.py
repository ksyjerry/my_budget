"""Regression #72 — members/export 컬럼명 표준화 (Empno → 사번)."""
import io
import pytest
from openpyxl import load_workbook
from sqlalchemy import text


@pytest.fixture(scope="function")
def export_seed(db):
    db.execute(text("DELETE FROM project_members WHERE project_code = 'AREA4-EXPORT-SRC'"))
    db.execute(text("DELETE FROM projects WHERE project_code = 'AREA4-EXPORT-SRC'"))
    db.commit()
    db.execute(text("""
        INSERT INTO projects (project_code, project_name, el_empno, pm_empno, template_status, contract_hours)
        VALUES ('AREA4-EXPORT-SRC', 'Export Test', '170661', '170661', '작성중', 100)
    """))
    db.execute(text("""
        INSERT INTO project_members (project_code, role, empno, name, grade, sort_order)
        VALUES ('AREA4-EXPORT-SRC', 'FLDT', '320915', '지해나', 'Senior', 1)
    """))
    db.commit()
    yield
    db.execute(text("DELETE FROM project_members WHERE project_code = 'AREA4-EXPORT-SRC'"))
    db.execute(text("DELETE FROM projects WHERE project_code = 'AREA4-EXPORT-SRC'"))
    db.commit()


def test_members_export_korean_column_headers(export_seed, client, elpm_cookie):
    resp = client.get("/api/v1/budget/projects/AREA4-EXPORT-SRC/members/export", cookies=elpm_cookie)
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    # 표준화: 사번 / 이름 / 역할 / 직급 / 팀(또는 부서) — "Empno" 컬럼명 제거
    assert "사번" in headers, f"'사번' 컬럼 없음. headers={headers}"
    assert "Empno" not in headers, f"Empno 컬럼이 남아있음. headers={headers}"
    assert "이름" in headers
    assert any(h in ("팀", "부서") for h in headers), f"팀/부서 컬럼 없음. headers={headers}"
