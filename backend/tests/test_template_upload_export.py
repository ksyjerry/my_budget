"""Tests for Step 3 template Excel export/upload (#20, #43)."""
import io

import pytest
from openpyxl import Workbook, load_workbook

from app.db.session import SessionLocal
from app.models.project import Project
from app.models.budget import BudgetDetail


TEST_PJ = "S5_TEST_TPL"


@pytest.fixture(autouse=True)
def _seed():
    s = SessionLocal()
    try:
        if s.query(Project).filter(Project.project_code == TEST_PJ).first() is None:
            s.add(Project(
                project_code=TEST_PJ,
                project_name="S5 Template Test",
                el_empno="170661", pm_empno="170661",
                contract_hours=100,
            ))
        s.query(BudgetDetail).filter(BudgetDetail.project_code == TEST_PJ).delete()
        s.add(BudgetDetail(
            project_code=TEST_PJ,
            budget_category="자산",
            budget_unit="매출채권-일반",
            empno="170661", emp_name="최성우",
            year_month="2025-04",
            budget_hours=10.0,
        ))
        s.commit()
    finally:
        s.close()
    yield


def test_template_export_xlsx(client, elpm_cookie):
    r = client.get(
        f"/api/v1/budget/projects/{TEST_PJ}/template/export",
        cookies=elpm_cookie,
    )
    assert r.status_code == 200, r.text
    assert "spreadsheetml.sheet" in r.headers.get("content-type", "")
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert "budget_category" in headers
    assert "budget_unit" in headers
    assert "empno" in headers
    # at least one month column
    has_month = any(h and "-" in str(h) and len(str(h)) >= 7 for h in headers if h)
    assert has_month


def test_template_upload_replaces_rows(client, elpm_cookie):
    wb = Workbook()
    ws = wb.active
    ws.append(["budget_category", "budget_unit", "empno", "name", "grade", "2025-04"])
    ws.append(["자산", "매출채권-일반", "170661", "최성우", "EL", 25.0])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    r = client.post(
        f"/api/v1/budget/projects/{TEST_PJ}/template/upload",
        files={"file": ("template.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        cookies=elpm_cookie,
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["imported_count"] >= 1


def test_staff_cannot_upload_template(client, staff_cookie):
    wb = Workbook()
    ws = wb.active
    ws.append(["budget_category", "budget_unit", "empno", "name", "grade", "2025-04"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post(
        f"/api/v1/budget/projects/{TEST_PJ}/template/upload",
        files={"file": ("template.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        cookies=staff_cookie,
    )
    assert r.status_code in (401, 403)
