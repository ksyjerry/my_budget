#!/usr/bin/env python3
"""Update Budget+ 2.0 Feedback_0425.xlsx with all Areas 1-6 answers (S7 final).

v2 — overrides v1 answers for items fixed in Areas 3-6 with actual completion status.
Generates fresh `_답변_v2.xlsx` from source xlsx.
"""
from __future__ import annotations
from pathlib import Path
import shutil
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


SRC = Path(__file__).parent.parent / "files" / "Budget+ 2.0 Feedback_0425.xlsx"
DST = Path(__file__).parent.parent / "files" / "Budget+ 2.0 Feedback_0425_답변_v2.xlsx"


# All row → answer mapping (Areas 1-6 complete state)
ANSWERS: dict[int, str] = {}


def add(row: int, body: str):
    ANSWERS[row] = body


# ─── Areas 1-2 결과 (이전 v1과 동일 — 변경 없음) ───────────────────────
# (#67-#71, #74, #99 — Area 1) (#79, #82, #84, #85, #120, #121 — Area 2) (#61, #98 — Area 2 POL-04)

add(63, """✅ 영역 2에서 처리 완료 (POL-04 표준형 워크플로우)

template_status 3단계 enum (작성중/작성완료/승인완료) + alembic 006 CHECK constraint.
신규 endpoints: POST /submit (PM), /approve (EL), /unlock (EL). 권한 가드 + Step 1/2/3 wizard에 워크플로우 버튼.

Provisional: POL-04 (b) — 김동환 외부 컨펌 대기.

PR: https://github.com/ksyjerry/my_budget/pull/2""")

add(65, """✓ 현행 유지 — 사용자 의견대로 추가 수정 불필요

잔금청구 완료된 프로젝트는 Azure SQL 마스터(BI_STAFFREPORT_PRJT_V) status 변경되어 자동 제외됨. 영역 6 진입 시 status 필터링 회귀 가드 추가.""")

add(66, """✅ 영역 2에서 처리 완료 (POL-05 (d) 하이브리드)

APScheduler _scheduled_tba_sync job — 매일 04:00 KST tracking._sync_tba_cache() 실행 + admin manual sync endpoint 유지.

Provisional: POL-05 (d) — 김미진 외부 컨펌 대기.

PR: https://github.com/ksyjerry/my_budget/pull/2""")

add(69, """✅ 영역 1에서 fix 완료 (회귀 #67)

docker-compose.yml command sh -c "npm run build && npm run start" + CI grep 가드 (scripts/ci/check-docker-compose-no-dev.sh) + frontend/tests/smoke/ 3 specs.

PR: https://github.com/ksyjerry/my_budget/pull/1""")

add(70, """✅ 영역 1에서 fix 완료 (회귀 #68)

QRP empno input readOnly 제거 + onChange 핸들러. Step 1에서 직접 입력 가능. regression test test_qrp_field_editable.spec.ts.

PR: https://github.com/ksyjerry/my_budget/pull/1 / Commit: f193434""")

add(71, """✅ 영역 1에서 fix 완료 (회귀 #69)

ProjectSearchModal: Azure SQL 비가용 시 PostgreSQL fallback + 클라이언트 미선택 상태 검색 가능. Frontend modal markup 개선 (role="dialog" + 검색 input).

PR: https://github.com/ksyjerry/my_budget/pull/1 / Commit: 4a8209a""")

add(72, """✅ 영역 1에서 fix 완료 (회귀 #70 + NumberField 추출)

NumberField → frontend/src/components/ui/NumberField.tsx 공유 컴포넌트. 안전 기본값 (min=0, allowNegative=false, displayThousandSeparator=true, step snap).
CI grep guard scripts/ci/check-no-direct-number-input.sh (Python multi-line) — 향후 NumberField 우회 차단.

PR: https://github.com/ksyjerry/my_budget/pull/1 / Commits: 8643db4, f4252c2""")

add(73, """✅ 영역 1에서 fix 완료 (회귀 #71)

EmployeeSearch.doSearch에 include_inactive=true + onKeyDown Enter 시 emp_status !== '재직' 검증 → alert + input clear.

PR: https://github.com/ksyjerry/my_budget/pull/1 / Commit: bdf4e7b""")

add(76, """✅ 영역 1에서 fix 완료 (회귀 #74)

NumberField step snap (Math.round(v/step)*step) — 0.24 → 0.25 자동 snap. min=0, max=300 default. regression test 모든 케이스 GREEN.

PR: https://github.com/ksyjerry/my_budget/pull/1 / Commits: 8643db4, 93fa80f""")

add(81, """✅ 영역 2에서 fix 완료

/projects/list 가시성 필터 EL_empno OR PM_empno + admin scope=all + anon → 빈 리스트. status 쿼리 파라미터 + updated_at 응답.
backend pytest test_list_endpoint_pm_visibility.py + frontend regression test_budget_list_states_visibility.spec.ts.

PR: https://github.com/ksyjerry/my_budget/pull/2 / Commit: df9f822""")

add(82, """(삭제 항목 — 답변 불필요)""")

add(84, """✅ 영역 2에서 fix 완료 — #79 동일 root cause

#79 fix로 PM 본인 프로젝트도 표시 → 임시저장 (작성중) 정상 조회. status filter 드롭다운 + 마지막 수정일 column 추가.

PR: https://github.com/ksyjerry/my_budget/pull/2 / Commits: df9f822, fe3d471, cc32f7d""")

add(86, """✅ 영역 2에서 fix 완료

목록 화면 "빈 Budget Template 다운로드" 버튼 제거. blank-export endpoint 자체는 유지 (Step 3 wizard 내부 사용).

PR: https://github.com/ksyjerry/my_budget/pull/2 / Commit: 649d40a""")

add(87, """✅ 영역 2에서 fix partial + 영역 5에서 보강

영역 2 (Task 13): 마지막 수정일 column 추가 (updated_at, ko-KR 날짜). 작성여부 필터 드롭다운 + 상태 배지 색상 차별화.

PR: https://github.com/ksyjerry/my_budget/pull/2""")

add(99, """(삭제 항목 — 답변 불필요)""")

add(101, """✅ 영역 1에서 fix 완료 (회귀 #99)

Step 1 헤더 div className 변경 (flex-wrap + gap-y-2 추가). 4개 버튼 boundingBox pairwise disjoint 검증. regression test test_step1_buttons_no_overlap.spec.ts.

PR: https://github.com/ksyjerry/my_budget/pull/1 / Commit: 868e2ba""")

add(122, """✅ 영역 2에서 fix 완료

Budget 입력 목록 화면 status filter dropdown (전체 / 작성중 / 작성완료 / 승인완료). 검색 + 필터 동시 적용. 상태별 색상 배지 (yellow/green/blue).

PR: https://github.com/ksyjerry/my_budget/pull/2 / Commit: cc32f7d""")

add(123, """✅ 영역 2에서 fix 완료 (회귀 #121)

클라이언트 사이드 필터 search.toLowerCase() 비교. EL/PM 이름도 검색 대상 추가. regression test_budget_list_search_case_insensitive.spec.ts.

PR: https://github.com/ksyjerry/my_budget/pull/2 / Commit: df57d51""")


# ─── Area 3 fixes (Step 1) ─────────────────────────────────────────────

add(59, """✅ 영역 3에서 fix 완료 (Step 1)

ClientSearchModal onSelect의 base.X || info.X || "" 패턴 → info.X || c.X || "" 로 변경. 이전 클라이언트 정보 자동 clear.
ProjectSearchModal에도 동일 stale-client 패턴 발견 → bonus fix 적용 (방어적).

PR: https://github.com/ksyjerry/my_budget/pull/3 / Commit: 26ba537""")

add(64, """✅ 영역 3에서 fix 완료 (Step 1)

Step1Form 섹션 순서 재배치: "프로젝트 정보"를 "클라이언트 기본정보" 위로 이동. 비감사 service_type 토글 시 사용자가 위쪽에서 작업 가능.

PR: https://github.com/ksyjerry/my_budget/pull/3 / Commit: e3ec7c2""")

add(88, """✅ 영역 3에서 fix 완료 (Step 1)

Backend /projects/{code}/clone-data: require_login 가드 추가 + project_name fallback (계속감사 케이스 회사명 검색 지원).
Frontend onCloneFromProject: credentials: "include" 추가 (login guard 호환).

PR: https://github.com/ksyjerry/my_budget/pull/3 / Commits: 1957c5a, b9e4490""")

add(102, """✅ 영역 3에서 fix 완료 (Step 1) — #86과 동일 fix

backend clone-data + frontend credentials 수정으로 더본코리아·롯데지알에스 두 케이스 모두 정상 작동.""")

add(103, """✅ 영역 3에서 fix 완료 (POL-06 (a) Step 1)

Step1Form에서 fulcrum_hours / ra_staff_hours / specialist_hours <NumberField> 3개 제거. 안내 문구 "Step 3 (Time Budget)에서 분배 입력합니다" 추가. state는 schema 유지 (Step 3에서 사용).

Provisional: POL-06 (a) — 홍상호 외부 컨펌 대기.

PR: https://github.com/ksyjerry/my_budget/pull/3 / Commit: 272f90f""")


# ─── Area 4 fixes (Step 2) ─────────────────────────────────────────────

add(74, """✅ 영역 4에서 fix 완료 (Step 2)

members/export 헤더 한글 표준화 ["사번", "이름", "역할", "직급", "팀"]. Empno 영어 컬럼명 제거. employees 마스터 join으로 팀 컬럼 채움.

PR: https://github.com/ksyjerry/my_budget/pull/4 / Commit: e31e625""")

add(75, """✅ 영역 4에서 fix 완료 (Step 2)

members/upload — 사번+이름만 필수, 직급/팀 누락 시 employees 마스터 lookup. 행 단위 errors[] 누적 (#87과 함께).

PR: https://github.com/ksyjerry/my_budget/pull/4 / Commit: e31e625""")

add(89, """✅ 영역 4에서 fix 완료 (Step 2)

members/upload 응답 schema {imported, skipped, errors: [{row, col, error}]}. 첫 실패에서 멈추지 않고 모든 오류 행 누적.

PR: https://github.com/ksyjerry/my_budget/pull/4 / Commit: e31e625""")

add(90, """✅ 영역 4에서 fix 완료 (Step 2)

employees/search 응답에 team_name 필드 추가 (department fallback). 프론트 EmployeeSearch dropdown에 팀명 column 표시.

PR: https://github.com/ksyjerry/my_budget/pull/4 / Commits: e31e625, bfd618c""")

add(104, """✅ 영역 4에서 fix 완료 (Step 2)

Step 2 FLDT 영역에 placeholder 멤버 추가 버튼 3개: "+ TBD" / "+ NS (New Step)" / "+ Associate". 클릭 시 placeholder 행 추가.

PR: https://github.com/ksyjerry/my_budget/pull/4 / Commit: bfd618c""")

add(105, """✅ 영역 4에서 fix 완료 (Step 2)

Step 2 지원 멤버 (Fulcrum/RA-Staff/Specialist) 자유 입력 → <select> dropdown 변경.

PR: https://github.com/ksyjerry/my_budget/pull/4 / Commit: bfd618c""")


# ─── Area 5 fixes (Step 3) ─────────────────────────────────────────────

add(60, """✅ 영역 5에서 fix 완료 (Step 3)

Step 3 grid 합계 행 colSpan 4 → 5 변경. 라벨이 cols 0-4 (checkbox/대분류/관리단위/담당자/직급) span, 합계 값이 col 5에 정렬.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commit: cc889df""")

add(77, """✅ 영역 5에서 fix 완료 (Step 3)

Step 3 export — 12개월 모두 컬럼 (값 0인 월도 빈 셀). budget_unit_master.sort_order 기준 정렬. db.expire_all() 로 fresh state read. upload merge 방식 (truncate 아님) — 비활성 행 보존.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commit: ff17a59""")

add(78, """✅ 영역 5에서 fix 완료 (Step 3)

비감사 service_type → AI suggest prompt에 service_type 별 ServiceTaskMaster 동적 조회 + 주입. 영역 5 alembic 007 (POL-03 (a)) subcategory_name 컬럼 활용.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commit: 646043f""")

add(85, """✅ 영역 5에서 fix 완료 (Step 3)

Excel template export 12개월 고정 컬럼 + sort_order 기준 정렬 + db.expire_all() 활성. upload merge 방식 채택. (#106/#107/#114/#117 통합 fix)

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commit: ff17a59""")

add(93, """✅ 영역 5에서 fix 완료 (Step 3)

Step 3 budget_details 에 unique constraint 부재 — 같은 unit에 여러 멤버 허용이 의도된 정책. 도큐먼트 추가 + reset endpoint로 일관 처리.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commit: 8e4a906""")

add(94, """✅ 영역 5에서 fix 완료 (Step 3) + alembic 007 (POL-03 (a))

service_task_master 에 subcategory_name 컬럼 추가 (alembic 007). 금융업 78행 시드 (#04 시트) — 소분류명 별도 저장.

Provisional: POL-03 (a) — 나형우/김지민 외부 컨펌 대기.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commits: 76af334, 65503a0""")

add(106, """✅ 영역 5에서 fix 완료 (Step 3)

Step 3 toolbar에 "📥 빈 Template 다운로드" 버튼 추가. backend blank-export 호출. #119 dropbox와 함께 설계.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commit: cc889df""")

add(107, """✅ 영역 5에서 fix 완료 (Step 3)

Excel export 가 항상 latest DB state 읽음 (db.expire_all()). Group A round-trip 일관성에 포함.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commit: ff17a59""")

add(108, """✅ 영역 5에서 fix 완료 (Step 3)

Excel export 가 12개월 모두 고정 컬럼 출력 (값 0인 월도 빈 셀). _build_12_months() helper.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commit: ff17a59""")

add(109, """✅ 영역 5에서 fix 완료 (Step 3)

Excel upload 를 truncate+insert → merge 방식. 업로드 파일에 있는 (budget_unit, empno) 만 갱신, 없는 행은 enabled=false 보존.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commit: ff17a59""")

add(110, """✅ 영역 5에서 fix 완료 (Step 3)

budget_assist suggest endpoint — system prompt에 initial_audit 컨텍스트 (계속감사 = 초도감사 0시간 / 초도감사 = 계획단계 비중 높게) 명시 주입.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commit: 646043f""")

add(111, """✅ 영역 5에서 fix 완료 (Step 3)

AI 추천 결과 NumberField default 천단위 표시 (영역 1 추상화 활용). 컬럼 width 조정.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commit: cc889df""")

add(112, """✅ 영역 5에서 fix 완료 (Step 3)

handleAiValidate에 aiAbortRef (useRef<AbortController>) — 새 호출 전 in-flight 요청 abort. AbortError silent swallow.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commit: cc889df""")

add(113, """✅ 영역 5에서 fix 완료 (Step 3)

backend error_sanitize.py 모듈 — IPv4 / localhost / URL 패턴 제거 후 사용자 표시. budget_assist 502/503 + budget_input upload 응답 적용. frontend sanitize() 안전망.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commits: aeca09f, cc889df""")

add(114, """✅ 영역 5에서 fix 완료 (Step 3)

Step 3 toolbar "전체 V 체크/해제" 토글 버튼. 1번 클릭으로 모든 row enabled flag 일괄 변경. 라벨 "체크" / "해제" 동적 변경.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commit: cc889df""")

add(115, """✅ 영역 5에서 fix 완료 (Step 3) — #112와 통합

전체 V 토글이 두 방향 모두 처리: 모두 체크 → "해제", 모두 해제 → "체크".""")

add(116, """✅ 영역 5에서 fix 완료 (Step 3)

Excel upload merge 방식 — 파일에 없는 budget_unit 은 enabled=false 로 비활성 유지 (삭제 안 함). 사용자 의도와 일치.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commit: ff17a59""")

add(117, """✅ 영역 5에서 fix 완료 (Step 3)

POST /projects/{code}/template/reset endpoint 추가 — backend budget_details 일괄 enabled=false + template_status 작성중 reset + change_log 기록. frontend 초기화 버튼이 backend reset 호출 후 frontend state도 reset.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commit: 8e4a906, cc889df""")

add(118, """✅ 영역 5에서 fix 완료 (Step 3) — #115와 통합""")

add(119, """✅ 영역 5에서 fix 완료 (Step 3)

Export + UI 모두 budget_unit_master.sort_order 기준 정렬. 화면 ↔ Excel 순서 일치.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commit: ff17a59""")

add(120, """✅ 영역 5에서 fix 완료 (Step 3) + alembic 007 (POL-07 (c))

projects.fiscal_end 컬럼 추가 (alembic 007). budget_service _build_12_months() 가 fiscal_end null 시 fiscal_start+11개월 default. Step 3 toolbar에 종료월 (<input type="month">) UI.

Provisional: POL-07 (c) — 신승엽 외부 컨펌 대기.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commits: 76af334, 65503a0, ff17a59, cc889df""")

add(121, """✅ 영역 5에서 fix 완료 (Step 3)

backend blank-export endpoint — openpyxl DataValidation + named range "BudgetUnitList" (hidden _lists 시트) 사용. Excel cell B2:B101에 dropdown 적용. Excel formula1 255-char limit 회피.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commit: cc889df""")


# ─── Area 6 fixes (Overview / Tracking) ────────────────────────────────

add(61, """✅ 영역 6에서 fix 완료 (Tracking + POL-08 (b))

Budget Tracking 화면 권한 가드 명확화: EL + admin (PM/Staff 차단). partner_access_config 와 정합. permission_matrix.yaml 에 tracking endpoints 추가 (영역 1 권한 매트릭스 확장).

Provisional: POL-08 (b) — 김동환 외부 컨펌 대기.

PR: https://github.com/ksyjerry/my_budget/pull/6""")

add(62, """✅ 영역 6에서 fix 완료 (Overview filter)

Overview 화면 모든 필터 dropdown — 검색 input 추가 (Project/EL/PM 등). 클라이언트 사이드 case-insensitive 필터.

PR: https://github.com/ksyjerry/my_budget/pull/6 / Commit: 9f8a498""")

add(67, """✅ 영역 6에서 fix 완료 (Details)

Backend assignments — emp_name 비어있으면 placeholder ("이름 미등록 ({empno})"). 타 LoS 인원 fallback. frontend 도 동일 매핑.

PR: https://github.com/ksyjerry/my_budget/pull/6 / Commit: 7655f6e""")

add(79, """✅ 영역 6에서 fix 완료 (Details + 분류)

filter-options endpoint — service_type display_label 매핑 ("AUDIT" → "감사" / 그 외 → "비감사"). frontend 분류 드롭다운에 통일된 표시.

PR: https://github.com/ksyjerry/my_budget/pull/6 / Commits: da3ecbd, 9f8a498""")

add(80, """✅ 영역 6에서 fix 완료 (Appendix)

xlsx 다운로드 응답 헤더 보강: Content-Type + RFC 5987 filename*=UTF-8''{quote(...)} (한글 파일명). 브라우저 보안 정책 호환.

PR: https://github.com/ksyjerry/my_budget/pull/6 / Commit: da3ecbd""")

add(95, """✅ 영역 6에서 fix 완료 (Overview QRP)

azure_service get_overview_actuals — qrp_empno 를 role_empnos 에 포함. QRP TMS 조회 정상 작동.

PR: https://github.com/ksyjerry/my_budget/pull/6 / Commit: 7655f6e""")

add(96, """✅ 영역 6에서 fix 완료 (Overview Staff Time)

staff_empnos 확장 (영역 1 S2 fix) 가 Overview 모든 view 에 적용 — Budget 없는 인원 actuals 도 집계.

PR: https://github.com/ksyjerry/my_budget/pull/6 / Commit: 7655f6e""")

add(97, """✅ 영역 6에서 fix 완료 (Overview filter)

연월 필터 dropdown 동적 생성 — 현재 회계연도 4월~3월 12개월 + 누적 옵션. My budget+ BI와 동일 UX.

PR: https://github.com/ksyjerry/my_budget/pull/6 / Commit: 9f8a498""")

add(98, """✅ 영역 6에서 fix 완료 (Overview filter cascading)

프로젝트 선택 → EL/PM dropdown filter cascading. selected project 의 el_empno / pm_empno 만 활성화.

PR: https://github.com/ksyjerry/my_budget/pull/6 / Commit: 9f8a498""")


# Override (#100 → Area 3 fix 동일하게 mark) — already done above


# ─── 사용자 테스트 필요 항목 (v2 — Area 1-6 모두 포함) ─────────────────

USER_TEST_ITEMS = [
    # Format: (No, 영역, 카테고리, 우선순위, 시나리오, 계정·데이터, 기대결과, 비고)
    # (이전 v1 27 항목 + Areas 3-6 추가 항목)

    # === Area 1 (회귀 7건) ===
    ("67", "영역 1", "배포 위생", "P0",
     "Docker prod build에서 dev overlay 0건 — /login, /overview-person, /budget-input 등 모든 페이지",
     "EL 계정",
     "nextjs-portal·data-nextjs-toast·data-nextjs-dialog 0개. console error 0",
     "PR #1"),

    ("68", "영역 1", "Step 1", "P1",
     "QRP 사번 입력 → 다른 필드 → QRP 다시 클릭 → 추가 입력",
     "EL 계정 (170661)",
     "값 유지 + 추가 입력 정상", "PR #1"),

    ("69", "영역 1", "Step 1", "P1",
     "신규 프로젝트 → 클라이언트 검색 skip → 프로젝트 검색만",
     "EL 또는 PM 계정",
     "비감사 service_type 프로젝트 결과에 표시", "PR #1"),

    ("70", "영역 1", "Step 1", "P1",
     "총 계약시간 12345 입력 → readOnly 필드 표시",
     "EL 계정",
     "ET 잔여시간에 '12,345' 천단위 콤마", "PR #1"),

    ("71", "영역 1", "Step 2", "P1",
     "휴직/퇴사 사번 입력 → Enter",
     "EL 계정 + 휴직 시드 사번",
     "alert 표시 + 등록 차단 + input clear", "PR #1, INACTIVE_EMPNO 시드 필요"),

    ("74", "영역 1", "Step 1/3", "P1",
     "(a) AX/DX -1 (b) Step 3 cell 0.24 (c) 301",
     "EL 계정",
     "(a) 0 clamp (b) 0.25 snap (c) 300 clamp", "PR #1"),

    ("99", "영역 1", "Step 1", "P1",
     "Step 1 모든 button 비겹침 — width 1024/1280/1920",
     "EL 계정",
     "어떤 width 에서도 button 겹침 없음", "PR #1"),

    # === Area 2 ===
    ("79+82", "영역 2", "Budget 목록", "P0",
     "PM 계정으로 /budget-input 접근",
     "PM 계정 (본인이 PM인 프로젝트 보유)",
     "본인 PM 프로젝트 모두 표시 (이전: EL만 보였음)", "PR #2"),

    ("84", "영역 2", "Budget 목록", "P2",
     "/budget-input 화면 상단",
     "EL 계정",
     "'빈 Budget Template 다운로드' 버튼 미노출", "PR #2"),

    ("85", "영역 2", "Budget 목록", "P2",
     "각 행에 '마지막 수정' 날짜 표시",
     "EL/PM 계정",
     "ko-KR 형식으로 표시", "PR #2"),

    ("120", "영역 2", "Budget 목록", "P1",
     "상태 필터 (전체/작성중/작성완료/승인완료) 선택",
     "EL 계정",
     "각 상태별 결과 즉시 변경", "PR #2"),

    ("121", "영역 2", "Budget 목록", "P1",
     "검색 'sk', 'SK', 'Sk텔레콤', 'SK텔레콤'",
     "EL 계정 (SK텔레콤 프로젝트 보유)",
     "모두 case-insensitive 매칭", "PR #2"),

    ("61/98", "영역 2", "POL-04 워크플로우", "P0",
     "PM submit → EL approve → unlock 전체 흐름 + Staff 직접 API 호출",
     "PM (170661) + EL + Staff",
     "각 단계 상태 변경 + Staff 차단", "PR #2"),

    ("64", "영역 2", "POL-05 daily sync", "P1",
     "Backend 재시작 → APScheduler 'daily_tba_sync' 등록 + 04:00 KST 자동 실행",
     "admin 계정",
     "Backend log 'Scheduler started' + 익일 신규 TBA 자동 추가", "PR #2"),

    # === Area 3 ===
    ("57", "영역 3", "Step 1", "P1",
     "클라이언트 A → B → C 변경 시퀀스",
     "EL 계정",
     "매번 의존 필드 (산업분류·자산규모) 새 client로 갱신", "PR #3"),

    ("62", "영역 3", "Step 1", "P2",
     "Step 1 화면 진입",
     "EL 계정",
     "프로젝트 정보 섹션이 클라이언트 정보 위에 위치", "PR #3"),

    ("86/100", "영역 3", "Step 1", "P1",
     "더본코리아 / 롯데지알에스 '이전 프로젝트 정보 가져오기'",
     "EL 계정 (시드 프로젝트 필요)",
     "alert + 시간/구성원/template 자동 채움", "PR #3, CLONE_SOURCE_PROJECT 시드 필요"),

    ("101", "영역 3", "Step 1", "P1",
     "Step 1 화면 — Fulcrum/RA-Staff/Specialist 라벨 0개. Step 3에서는 여전히 존재",
     "EL 계정",
     "Step 1 입력칸 제거 (POL-06 (a))", "PR #3"),

    # === Area 4 ===
    ("72", "영역 4", "Step 2", "P1",
     "Step 2 Excel export 다운",
     "EL 계정",
     "헤더 '사번/이름/역할/직급/팀' (한글)", "PR #4"),

    ("73+87", "영역 4", "Step 2", "P1",
     "Excel 사번+이름만 입력 후 업로드 / 결함 행 다수 업로드",
     "EL 계정",
     "사번+이름만 OK + 모든 결함 행 누적 응답", "PR #4"),

    ("88", "영역 4", "Step 2", "P1",
     "EmployeeSearch 결과 dropdown",
     "EL 계정",
     "팀명 column 표시 (동명이인 구분)", "PR #4"),

    ("102", "영역 4", "Step 2", "P2",
     "'+ TBD' / '+ NS' / '+ Associate' 버튼 클릭",
     "EL 계정",
     "placeholder 행 추가됨", "PR #4"),

    ("103", "영역 4", "Step 2", "P2",
     "Fulcrum/RA-Staff/Specialist 입력",
     "EL 계정",
     "<select> dropdown — 자유 텍스트 불가", "PR #4"),

    # === Area 5 (대규모) ===
    ("Group A", "영역 5", "Excel I/O", "P0",
     "저장 후 다운로드 / 초기화 후 다운로드 / 12개월 컬럼 / merge upload / 화면-Excel 순서 일치",
     "EL 계정",
     "round-trip 일관성", "PR #5 Group A"),

    ("Group B+G", "영역 5", "초기화 + 정책", "P1",
     "초기화 → step 이동 → 재진입 깨끗 + 종료월 입력 → 저장 → 재로드 유지",
     "EL 계정",
     "frontend+backend reset 일관 + fiscal_end 유지", "PR #5"),

    ("Group C", "영역 5", "Step 3 UX", "P2",
     "전체 V 토글 버튼 — 체크/해제 양방향",
     "EL 계정",
     "1번 클릭으로 모두 체크/해제", "PR #5"),

    ("Group D", "영역 5", "Step 3 layout", "P2",
     "Step 3 합계 칸 정렬 확인",
     "EL 계정",
     "정상 정렬 (col 5 합계 값)", "PR #5"),

    ("Group E", "영역 5", "AI Assist", "P1",
     "비감사 service_type 추천 / 계속감사 → 초도감사 0 / cancel handling / 등록오류 IP 미노출",
     "EL 계정 (LLM API 작동 환경)",
     "service_type 매핑 + 컨텍스트 정확", "PR #5"),

    ("Group F", "영역 5", "금융업 시드 + POL-03", "P0",
     "감사 프로젝트 → Step 3 에 금융업 78행 budget_unit 노출",
     "EL 계정",
     "대출채권/보험계약부채/KICS 등 표시 + subcategory_name 분리", "PR #5"),

    ("Group H", "영역 5", "Template", "P1",
     "Step 3 '빈 Template 다운로드' → Excel data validation list",
     "EL 계정",
     "cell 클릭 시 dropdown 표시", "PR #5"),

    # === Area 6 ===
    ("POL-01 (b)", "영역 6", "Budget 정의", "P0",
     "Overview KPI / Project 테이블 / Tracking / Summary 모두 Budget = 총계약 - AX/DX 확인",
     "EL 계정",
     "모든 view 동일 정의 + 실질 Progress 컬럼", "PR #6"),

    ("65", "영역 6", "Details", "P2",
     "타 LoS 인원 — 이름 미등록 시 fallback 표시",
     "EL 계정 (시드 필요)",
     "'이름 미등록 ({empno})' 표시", "PR #6"),

    ("77", "영역 6", "분류", "P2",
     "service_type 표시 — '감사/비감사' 통일",
     "EL 계정",
     "기타 service_type 도 '비감사' 그룹", "PR #6"),

    ("78", "영역 6", "Appendix", "P2",
     "xlsx 다운로드 — 한글 파일명 포함",
     "EL 계정",
     "다운로드 차단 경고 미발생", "PR #6"),

    ("93", "영역 6", "Overview", "P1",
     "QRP empno 있는 프로젝트 — Actual 표시",
     "EL 계정 (시드 + Azure SQL 가용)",
     "QRP TMS 조회 정상", "PR #6"),

    ("94", "영역 6", "Overview", "P1",
     "Budget 없는 인원 actuals — Overview Project 테이블",
     "EL 계정 + TMS data",
     "actuals 집계 포함", "PR #6"),

    ("95", "영역 6", "Overview filter", "P1",
     "연월 dropdown — 4월~3월 12개월",
     "EL 계정",
     "월별 선택 가능 + 누적 옵션", "PR #6"),

    ("96", "영역 6", "Overview filter", "P1",
     "프로젝트 선택 → EL/PM dropdown",
     "EL 계정",
     "해당 프로젝트의 EL/PM 만 활성화 (cascading)", "PR #6"),

    ("POL-08 (b)", "영역 6", "Budget Tracking 권한", "P0",
     "EL/admin/PM/Staff 각각 Tracking 화면 접근",
     "각 페르소나 계정",
     "EL+admin OK / PM+Staff 403", "PR #6"),

    # === 공통 ===
    ("CI 게이트", "영역 1+2 공통", "CI 게이트", "P0",
     "(a) <input type=number> 추가 PR (b) 게이트 미통과 PR merge 시도",
     "GitHub repo write 권한",
     "(a) Grep Guards 실패 (b) merge 차단", "Branch protection 적용 후"),

    ("Layer 3 — 모든 Area", "영역 1-6 종료", "사용자 컨펌", "P0",
     "각 Area 의 qa-checklist 일괄 staging 검증",
     "EL + PM + Staff + admin",
     "모든 결함 차단/fix 유지 + qa-checklist all PASS",
     "S7 sign-off → main merge"),

    # === POL 외부 결정자 컨펌 ===
    ("POL-01", "정책", "외부 결정자", "P0",
     "Budget 정의 (b) 총계약-AX/DX confirm",
     "김동환 (4/27 회의)",
     "정식 확정 → policy-decisions.md 업데이트", "영역 6 적용"),

    ("POL-02", "정책", "외부 결정자", "P0",
     "통상자문/내부회계 (b) service_type별 다름",
     "신승엽 (4/27)",
     "정식 확정 + 통상자문 UI mini-cycle 트리거", "영역 5 schema 준비됨"),

    ("POL-03", "정책", "외부 결정자", "P0",
     "비감사 관리단위 (a) 소분류명 별도 컬럼",
     "나형우/김지민 (4/27)",
     "정식 확정 → 영역 5 시드 패턴 유지", "alembic 007 적용됨"),

    ("POL-04", "정책", "외부 결정자 (provisional)", "P1",
     "EL 승인 워크플로우 (b) 표준형",
     "김동환",
     "정식 확정 OR (a)/(c) 결정 시 영역 2 후속 fix", "영역 2 적용"),

    ("POL-05", "정책", "외부 결정자 (provisional)", "P2",
     "TBA sync (d) 하이브리드 (Daily 04:00 + manual)",
     "김미진",
     "정식 확정 OR 다른 안 시 영역 2 후속", "영역 2 적용"),

    ("POL-06", "정책", "외부 결정자 (provisional)", "P0",
     "RA 주관 FLDT (a) Step 1 제거 / Step 3에만",
     "홍상호",
     "정식 확정 OR 다른 안 시 영역 3 후속", "영역 3 적용"),

    ("POL-07", "정책", "외부 결정자 (provisional)", "P1",
     "프로젝트 기간 (c) 시작 자동 + 끝 수동",
     "신승엽",
     "정식 확정 OR 다른 안 시 영역 5 후속", "alembic 007 + 영역 5 적용"),

    ("POL-08", "정책", "외부 결정자 (provisional)", "P1",
     "Budget Tracking 권한 (b) EL+admin",
     "김동환",
     "정식 확정 OR PM 노출 결정 시 영역 6 후속", "영역 6 적용"),

    ("POL-09", "정책", "내부 결정", "P3",
     "TBA sync 권한 admin only",
     "owner",
     "admin only 권장 (YAGNI)", "영역 2 회고 포함"),
]


def fill_answers(ws):
    align = Alignment(wrap_text=True, vertical="top")
    border = Side(style="thin", color="DDDDDD")
    cell_border = Border(top=border, bottom=border, left=border, right=border)
    for row, body in ANSWERS.items():
        c = ws.cell(row=row, column=7, value=body)
        c.alignment = align
        c.border = cell_border
        ws.row_dimensions[row].height = max(60, min(400, body.count("\n") * 14 + 40))
    ws.column_dimensions["G"].width = 80


def add_user_test_sheet(wb):
    if "사용자_테스트_필요_항목" in wb.sheetnames:
        del wb["사용자_테스트_필요_항목"]
    ws = wb.create_sheet("사용자_테스트_필요_항목", 1)

    headers = ["피드백 No", "영역", "카테고리", "우선순위",
               "테스트 시나리오", "사용 계정·데이터", "기대 결과", "비고"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="D04A02")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Side(style="thin", color="888888")
    cell_border = Border(top=border, bottom=border, left=border, right=border)

    ws.cell(row=1, column=1, value=(
        f"My Budget+ 2.0 — 사용자 테스트 필요 항목 (S7 완료, v2 작성: {datetime.now().strftime('%Y-%m-%d')}) | "
        f"6개 영역 (PR #1-#6) staging 검증 + 9개 POL 외부 결정자 컨펌"
    ))
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.row_dimensions[1].height = 36

    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = cell_border
    ws.row_dimensions[3].height = 32

    align_top = Alignment(wrap_text=True, vertical="top")
    align_top_center = Alignment(wrap_text=True, vertical="top", horizontal="center")
    priority_color = {
        "P0": PatternFill("solid", fgColor="FFCDD2"),
        "P1": PatternFill("solid", fgColor="FFE0B2"),
        "P2": PatternFill("solid", fgColor="FFF9C4"),
        "P3": PatternFill("solid", fgColor="C8E6C9"),
    }
    for r_idx, item in enumerate(USER_TEST_ITEMS, 4):
        for c_idx, val in enumerate(item, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.border = cell_border
            if c_idx == 4:
                c.alignment = align_top_center
                c.fill = priority_color.get(val, PatternFill())
            else:
                c.alignment = align_top
        max_len = max(len(str(v)) for v in item)
        ws.row_dimensions[r_idx].height = max(40, min(180, (max_len // 50 + 1) * 28))

    widths = {"A": 18, "B": 14, "C": 22, "D": 10, "E": 50, "F": 30, "G": 50, "H": 30}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A4"


def main():
    if not SRC.exists():
        raise SystemExit(f"Source file missing: {SRC}")
    shutil.copy(SRC, DST)
    wb = openpyxl.load_workbook(DST)

    if "#01_사용자 feedback" not in wb.sheetnames:
        raise SystemExit("Sheet '#01_사용자 feedback' missing")

    fill_answers(wb["#01_사용자 feedback"])
    add_user_test_sheet(wb)

    wb.save(DST)
    print(f"Updated: {DST}")
    print(f"  답변 작성: {len(ANSWERS)} 행")
    print(f"  신규 시트: 사용자_테스트_필요_항목 ({len(USER_TEST_ITEMS)} 항목)")


if __name__ == "__main__":
    main()
