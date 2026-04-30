#!/usr/bin/env python3
"""Update Budget+ 2.0 Feedback_0425.xlsx with S7 + S8 답변 (v3 — 2026-04-29).

v3 — overrides v2 answers for items affected by:
- S8: 사용자용 Excel I/O 전면 제거 (업로드/다운로드/Appendix 페이지)
- S8: Step 3 Grid UX 4건 (sticky / 분배 도우미 / 실시간 검증 / 검색·접기)
- S8 후속: Step 1 섹션 순서 재배치 (서비스→클라이언트→프로젝트→시간), 프로젝트 검색 client_code→client_name 매핑 fix
- S8 후속: V 토글 그리드 헤더로 이동 (Toolbar 버튼 → V 열 마스터 체크박스)

Output: files/Budget+ 2.0 Feedback_0425_답변_v3.xlsx
"""
from __future__ import annotations
from pathlib import Path
import shutil
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


SRC = Path(__file__).parent.parent / "files" / "Budget+ 2.0 Feedback_0425.xlsx"
DST = Path(__file__).parent.parent / "files" / "Budget+ 2.0 Feedback_0425_답변_v3.xlsx"


# ─── v2 baseline ANSWERS (원본 그대로 유지) ───────────────────────────────
ANSWERS: dict[int, str] = {}


def add(row: int, body: str):
    ANSWERS[row] = body


# v2 의 모든 답변을 그대로 import — 동일한 dict 채움
add(63, """✅ 영역 2에서 처리 완료 (POL-04 표준형 워크플로우)

template_status 3단계 enum (작성중/작성완료/승인완료) + alembic 006 CHECK constraint.
신규 endpoints: POST /submit (PM), /approve (EL), /unlock (EL). 권한 가드 + Step 1/2/3 wizard에 워크플로우 버튼.

Provisional: POL-04 (b) — 김동환 외부 컨펌 대기.

PR: https://github.com/ksyjerry/my_budget/pull/2""")

add(65, """✓ 현행 유지 — 사용자 의견대로 추가 수정 불필요

잔금청구 완료된 프로젝트는 Azure SQL 마스터(BI_STAFFREPORT_PRJT_V) status 변경되어 자동 제외됨. 영역 6 진입 시 status 필터링 회귀 가드 추가.""")

add(66, """✅ 영역 2에서 처리 완료 (POL-05 (d) 하이브리드)

APScheduler _scheduled_tba_sync job — 매일 04:00 KST tracking._sync_tba_cache() 실행 + admin manual sync endpoint 유지.

Provisional: POL-05 (d) — 김미진 외부 컨펌 대기.

PR: https://github.com/ksyjerry/my_budget/pull/2""")

add(69, """✅ 영역 1에서 fix 완료 (회귀 #67)

docker-compose.yml command sh -c "npm run build && npm run start" + CI grep 가드 (scripts/ci/check-docker-compose-no-dev.sh) + frontend/tests/smoke/ 3 specs.

PR: https://github.com/ksyjerry/my_budget/pull/1""")

add(70, """✅ 영역 1에서 fix 완료 (회귀 #68)

QRP empno input readOnly 제거 + onChange 핸들러. Step 1에서 직접 입력 가능. regression test test_qrp_field_editable.spec.ts.

PR: https://github.com/ksyjerry/my_budget/pull/1 / Commit: f193434""")

add(72, """✅ 영역 1에서 fix 완료 (회귀 #70 + NumberField 추출)

NumberField → frontend/src/components/ui/NumberField.tsx 공유 컴포넌트. 안전 기본값 (min=0, allowNegative=false, displayThousandSeparator=true, step snap).
CI grep guard scripts/ci/check-no-direct-number-input.sh (Python multi-line) — 향후 NumberField 우회 차단.

PR: https://github.com/ksyjerry/my_budget/pull/1 / Commits: 8643db4, f4252c2""")

add(73, """✅ 영역 1에서 fix 완료 (회귀 #71)

EmployeeSearch.doSearch에 include_inactive=true + onKeyDown Enter 시 emp_status !== '재직' 검증 → alert + input clear.

PR: https://github.com/ksyjerry/my_budget/pull/1 / Commit: bdf4e7b""")

add(76, """✅ 영역 1에서 fix 완료 (회귀 #74)

NumberField step snap (Math.round(v/step)*step) — 0.24 → 0.25 자동 snap. min=0, max=300 default. regression test 모든 케이스 GREEN.

PR: https://github.com/ksyjerry/my_budget/pull/1 / Commits: 8643db4, 93fa80f""")

add(81, """✅ 영역 2에서 fix 완료

/projects/list 가시성 필터 EL_empno OR PM_empno + admin scope=all + anon → 빈 리스트. status 쿼리 파라미터 + updated_at 응답.
backend pytest test_list_endpoint_pm_visibility.py + frontend regression test_budget_list_states_visibility.spec.ts.

PR: https://github.com/ksyjerry/my_budget/pull/2 / Commit: df9f822""")

add(82, """(삭제 항목 — 답변 불필요)""")

add(84, """✅ 영역 2에서 fix 완료 — #79 동일 root cause

#79 fix로 PM 본인 프로젝트도 표시 → 임시저장 (작성중) 정상 조회. status filter 드롭다운 + 마지막 수정일 column 추가.

PR: https://github.com/ksyjerry/my_budget/pull/2 / Commits: df9f822, fe3d471, cc32f7d""")

add(86, """✅ 영역 2에서 fix 완료 — S8 에서 한 번 더 정리

영역 2: 목록 화면 "빈 Budget Template 다운로드" 버튼 제거.
S8 (PR #9): blank-export endpoint 자체도 제거 (사용자용 Excel 기능 전면 제거 정책).

PR: https://github.com/ksyjerry/my_budget/pull/2 / Commit: 649d40a
PR: https://github.com/ksyjerry/my_budget/pull/9 (S8 — Excel I/O 제거)""")

add(87, """✅ 영역 2에서 fix partial + 영역 5에서 보강

영역 2 (Task 13): 마지막 수정일 column 추가 (updated_at, ko-KR 날짜). 작성여부 필터 드롭다운 + 상태 배지 색상 차별화.

PR: https://github.com/ksyjerry/my_budget/pull/2""")

add(99, """(삭제 항목 — 답변 불필요)""")

add(101, """✅ 영역 1에서 fix 완료 (회귀 #99)

Step 1 헤더 div className 변경 (flex-wrap + gap-y-2 추가). 4개 버튼 boundingBox pairwise disjoint 검증. regression test test_step1_buttons_no_overlap.spec.ts.

PR: https://github.com/ksyjerry/my_budget/pull/1 / Commit: 868e2ba""")

add(122, """✅ 영역 2에서 fix 완료

Budget 입력 목록 화면 status filter dropdown (전체 / 작성중 / 작성완료 / 승인완료). 검색 + 필터 동시 적용. 상태별 색상 배지 (yellow/green/blue).

PR: https://github.com/ksyjerry/my_budget/pull/2 / Commit: cc32f7d""")

add(123, """✅ 영역 2에서 fix 완료 (회귀 #121)

클라이언트 사이드 필터 search.toLowerCase() 비교. EL/PM 이름도 검색 대상 추가. regression test_budget_list_search_case_insensitive.spec.ts.

PR: https://github.com/ksyjerry/my_budget/pull/2 / Commit: df57d51""")


# ─── Area 3 fixes (Step 1) ─────────────────────────────────────────────

add(59, """✅ 영역 3에서 fix 완료 (Step 1)

ClientSearchModal onSelect의 base.X || info.X || "" 패턴 → info.X || c.X || "" 로 변경. 이전 클라이언트 정보 자동 clear.
ProjectSearchModal에도 동일 stale-client 패턴 발견 → bonus fix 적용 (방어적).

PR: https://github.com/ksyjerry/my_budget/pull/3 / Commit: 26ba537""")

# row 64 → S8 override (아래 S8_OVERRIDES 에서 갱신)

add(88, """✅ 영역 3에서 fix 완료 (Step 1)

Backend /projects/{code}/clone-data: require_login 가드 추가 + project_name fallback (계속감사 케이스 회사명 검색 지원).
Frontend onCloneFromProject: credentials: "include" 추가 (login guard 호환).

PR: https://github.com/ksyjerry/my_budget/pull/3 / Commits: 1957c5a, b9e4490""")

add(102, """✅ 영역 3에서 fix 완료 (Step 1) — #86과 동일 fix

backend clone-data + frontend credentials 수정으로 더본코리아·롯데지알에스 두 케이스 모두 정상 작동.""")

add(103, """✅ 영역 3에서 fix 완료 (POL-06 (a) Step 1)

Step1Form에서 fulcrum_hours / ra_staff_hours / specialist_hours <NumberField> 3개 제거. 안내 문구 "Step 3 (Time Budget)에서 분배 입력합니다" 추가. state는 schema 유지 (Step 3에서 사용).

Provisional: POL-06 (a) — 홍상호 외부 컨펌 대기.

PR: https://github.com/ksyjerry/my_budget/pull/3 / Commit: 272f90f""")


# ─── Area 4 fixes (Step 2) ─────────────────────────────────────────────
# row 74, 75, 89, 90 → S8 override (Excel 제거)

add(90, """✅ 영역 4에서 fix 완료 (Step 2)

employees/search 응답에 team_name 필드 추가 (department fallback). 프론트 EmployeeSearch dropdown에 팀명 column 표시.

PR: https://github.com/ksyjerry/my_budget/pull/4 / Commits: e31e625, bfd618c""")

add(104, """✅ 영역 4에서 fix 완료 (Step 2)

Step 2 FLDT 영역에 placeholder 멤버 추가 버튼 3개: "+ TBD" / "+ NS (New Step)" / "+ Associate". 클릭 시 placeholder 행 추가.

PR: https://github.com/ksyjerry/my_budget/pull/4 / Commit: bfd618c""")

add(105, """✅ 영역 4에서 fix 완료 (Step 2)

Step 2 지원 멤버 (Fulcrum/RA-Staff/Specialist) 자유 입력 → <select> dropdown 변경.

PR: https://github.com/ksyjerry/my_budget/pull/4 / Commit: bfd618c""")


# ─── Area 5 fixes (Step 3) ─────────────────────────────────────────────

add(60, """✅ 영역 5에서 fix 완료 + S8에서 sticky 헤더 강화

영역 5: Step 3 grid 합계 행 colSpan 4 → 5 변경. 라벨이 cols 0-4 (checkbox/대분류/관리단위/담당자/직급) span, 합계 값이 col 5에 정렬.
S8 (PR #9): position: sticky + border-collapse: separate 적용. 헤더/좌측 4 column/우측 합계 column 모두 sticky — 스크롤해도 정렬 흐트러짐 없음.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commit: cc889df
PR: https://github.com/ksyjerry/my_budget/pull/9 / Commit: 9c34f41 (sticky)""")

# row 77, 85, 106-109, 116, 119, 121 → S8 override (Excel 제거)

add(78, """✅ 영역 5에서 fix 완료 (Step 3)

비감사 service_type → AI suggest prompt에 service_type 별 ServiceTaskMaster 동적 조회 + 주입. 영역 5 alembic 007 (POL-03 (a)) subcategory_name 컬럼 활용.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commit: 646043f""")

add(93, """✅ 영역 5에서 fix 완료 (Step 3)

Step 3 budget_details 에 unique constraint 부재 — 같은 unit에 여러 멤버 허용이 의도된 정책. 도큐먼트 추가 + reset endpoint로 일관 처리.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commit: 8e4a906""")

add(94, """✅ 영역 5에서 fix 완료 (Step 3) + alembic 007 (POL-03 (a))

service_task_master 에 subcategory_name 컬럼 추가 (alembic 007). 금융업 78행 시드 (#04 시트) — 소분류명 별도 저장.

Provisional: POL-03 (a) — 나형우/김지민 외부 컨펌 대기.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commits: 76af334, 65503a0""")

add(110, """✅ 영역 5에서 fix 완료 (Step 3)

budget_assist suggest endpoint — system prompt에 initial_audit 컨텍스트 (계속감사 = 초도감사 0시간 / 초도감사 = 계획단계 비중 높게) 명시 주입.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commit: 646043f""")

add(111, """✅ 영역 5에서 fix 완료 (Step 3)

AI 추천 결과 NumberField default 천단위 표시 (영역 1 추상화 활용). 컬럼 width 조정.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commit: cc889df""")

add(112, """✅ 영역 5에서 fix 완료 (Step 3)

handleAiValidate에 aiAbortRef (useRef<AbortController>) — 새 호출 전 in-flight 요청 abort. AbortError silent swallow.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commit: cc889df""")

add(113, """✅ 영역 5에서 fix 완료 (Step 3)

backend error_sanitize.py 모듈 — IPv4 / localhost / URL 패턴 제거 후 사용자 표시. budget_assist 502/503 + budget_input upload 응답 적용. frontend sanitize() 안전망.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commits: aeca09f, cc889df""")

# row 114, 115 → S8 override (그리드 헤더 마스터 체크박스로 이동)

add(117, """✅ 영역 5에서 fix 완료 (Step 3)

POST /projects/{code}/template/reset endpoint 추가 — backend budget_details 일괄 enabled=false + template_status 작성중 reset + change_log 기록. frontend 초기화 버튼이 backend reset 호출 후 frontend state도 reset.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commit: 8e4a906, cc889df""")

add(118, """✅ 영역 5에서 fix 완료 (Step 3) — #115와 통합""")

add(120, """✅ 영역 5에서 fix 완료 (Step 3) + alembic 007 (POL-07 (c))

projects.fiscal_end 컬럼 추가 (alembic 007). budget_service _build_12_months() 가 fiscal_end null 시 fiscal_start+11개월 default. Step 3 toolbar에 종료월 (<input type="month">) UI.

Provisional: POL-07 (c) — 신승엽 외부 컨펌 대기.

PR: https://github.com/ksyjerry/my_budget/pull/5 / Commits: 76af334, 65503a0, ff17a59, cc889df""")


# ─── Area 6 fixes (Overview / Tracking) ────────────────────────────────

add(61, """✅ 영역 6에서 fix 완료 (Tracking + POL-08 (b))

Budget Tracking 화면 권한 가드 명확화: EL + admin (PM/Staff 차단). partner_access_config 와 정합. permission_matrix.yaml 에 tracking endpoints 추가 (영역 1 권한 매트릭스 확장).

Provisional: POL-08 (b) — 김동환 외부 컨펌 대기.

PR: https://github.com/ksyjerry/my_budget/pull/6""")

add(62, """✅ 영역 6에서 fix 완료 (Overview filter)

Overview 화면 모든 필터 dropdown — 검색 input 추가 (Project/EL/PM 등). 클라이언트 사이드 case-insensitive 필터.

PR: https://github.com/ksyjerry/my_budget/pull/6 / Commit: 9f8a498""")

add(67, """✅ 영역 6에서 fix 완료 (Details)

Backend assignments — emp_name 비어있으면 placeholder ("이름 미등록 ({empno})"). 타 LoS 인원 fallback. frontend 도 동일 매핑.

PR: https://github.com/ksyjerry/my_budget/pull/6 / Commit: 7655f6e""")

add(79, """✅ 영역 6에서 fix 완료 (Details + 분류)

filter-options endpoint — service_type display_label 매핑 ("AUDIT" → "감사" / 그 외 → "비감사"). frontend 분류 드롭다운에 통일된 표시.

PR: https://github.com/ksyjerry/my_budget/pull/6 / Commits: da3ecbd, 9f8a498""")

# row 80 → S8 override (Appendix 자체 제거)

add(95, """✅ 영역 6에서 fix 완료 (Overview QRP)

azure_service get_overview_actuals — qrp_empno 를 role_empnos 에 포함. QRP TMS 조회 정상 작동.

PR: https://github.com/ksyjerry/my_budget/pull/6 / Commit: 7655f6e""")

add(96, """✅ 영역 6에서 fix 완료 (Overview Staff Time)

staff_empnos 확장 (영역 1 S2 fix) 가 Overview 모든 view 에 적용 — Budget 없는 인원 actuals 도 집계.

PR: https://github.com/ksyjerry/my_budget/pull/6 / Commit: 7655f6e""")

add(97, """✅ 영역 6에서 fix 완료 (Overview filter)

연월 필터 dropdown 동적 생성 — 현재 회계연도 4월~3월 12개월 + 누적 옵션. My budget+ BI와 동일 UX.

PR: https://github.com/ksyjerry/my_budget/pull/6 / Commit: 9f8a498""")

add(98, """✅ 영역 6에서 fix 완료 (Overview filter cascading)

프로젝트 선택 → EL/PM dropdown filter cascading. selected project 의 el_empno / pm_empno 만 활성화.

PR: https://github.com/ksyjerry/my_budget/pull/6 / Commit: 9f8a498""")


# ────────────────────────────────────────────────────────────────────────
# ─── S8 OVERRIDES (2026-04-29 이후 추가/변경) ─────────────────────────────
# ────────────────────────────────────────────────────────────────────────

# Excel 관련 — 사용자 가시 기능 전면 제거 (S8 정책)
EXCEL_REMOVED_NOTE = """🗑 사용자용 Excel 기능 제거됨 (S8 — 2026-04-29)

S8 사이클에서 다음 사용자용 Excel I/O 모두 제거:
- Backend: excel_parser, excel_export, budget_upload.py, export.py 모듈 삭제
- Backend: budget_input.py 의 Excel endpoints (members/template export·upload, blank-export) 삭제
- Frontend: Appendix 페이지 + 사이드바 메뉴 + Step 2/3 Excel 버튼 제거
- 약 -1,830 LOC 순감

대신 Step 3 그리드 UX 강화 (S8 신규):
- Sticky 헤더 + 좌측 4 column + 우측 합계 column
- 분배 도우미 (균등 / 기말 / 유사회사 비율)
- 실시간 검증 + cell highlighting + 진행률 bar
- 검색 input + 대분류 접기/펼치기 + sessionStorage 영속성

PR: https://github.com/ksyjerry/my_budget/pull/9"""


# row 22 (#21) — 표감 처럼 엑셀 업/다운 추가
add(22, EXCEL_REMOVED_NOTE + """

원래 답변 (영역 5): Step 3 toolbar Excel 다운/업로드 버튼 + GET/POST endpoint 추가.""")

# row 34 (#33) — Appendix 프로젝트별 다운로드
add(34, """🗑 Appendix 페이지 자체 제거됨 (S8 — 2026-04-29)

Appendix 페이지의 모든 Excel 다운로드 기능이 사용자 요구로 제거 — 사이드바 메뉴도 제거됨.
원래 fix (영역 6): 프로젝트 dropdown 추가하여 선택적 다운로드.

PR: https://github.com/ksyjerry/my_budget/pull/9""")

# row 35 (#34) — 빈 Budget Template 다운로드
add(35, EXCEL_REMOVED_NOTE + """

원래 답변 (영역 5+6): blank-export endpoint + Budget 입력 목록 페이지 빈 Template 버튼.""")

# row 42 (#39) — Step 2 Excel import/export
add(42, EXCEL_REMOVED_NOTE + """

원래 답변 (영역 4): /members/export, /members/upload + Step 2 상단 Excel 버튼.""")

# row 45 (#41) — Step 3 Excel import/export
add(45, EXCEL_REMOVED_NOTE + """

원래 답변 (영역 5): Step 3 Time Budget Excel 업/다운로드 + round-trip 일관성.""")

# row 47 (#43) — Appendix CSV→XLSX
add(47, """🗑 Appendix 페이지 자체 제거됨 (S8 — 2026-04-29)

원래 fix (영역 6): Appendix 페이지 안내 문구 "CSV → XLSX" 변경 + 실제 다운로드도 xlsx 통일.

PR: https://github.com/ksyjerry/my_budget/pull/9""")

# row 64 (#61) — 비감사 클릭 시 프로젝트 정보 상단 (이번 회차에서 추가 처리)
add(64, """✅ 영역 3 + S8 후속 fix 완료 (Step 1 레이아웃 재배치)

영역 3 (PR #3): Step1Form 섹션 순서 재배치 — "프로젝트 정보"를 "클라이언트 기본정보" 위로 이동.

S8 후속 fix (PR #9, 2026-04-29): 사용자 추가 요청으로 최종 순서 재정렬:
1. 서비스 분류 (별도 최상단 섹션으로 분리)
2. 클라이언트 기본정보
3. 프로젝트 정보
4. 시간 배분

→ 비감사 service_type 토글이 위쪽에 위치 + 클라이언트 정보가 프로젝트보다 먼저 나옴.

PR: https://github.com/ksyjerry/my_budget/pull/3 / Commit: e3ec7c2 (영역 3)
PR: https://github.com/ksyjerry/my_budget/pull/9 / Commit: 2dc48d1 (S8 후속)""")

# row 71 (#68) — 클라이언트 미선택 프로젝트 검색 (이번 회차 client_code → client_name 매핑 fix 추가)
add(71, """✅ 영역 1 fix + S8 후속 client_code → client_name 매핑 fix 완료

영역 1 (PR #1): ProjectSearchModal Azure SQL 비가용 시 PG fallback + 클라이언트 미선택 검색 + modal markup 개선.

S8 후속 fix (PR #9, 2026-04-29): legal entity client_code (예: 05319 삼성전자) ≠ Azure project_code 앞 5자리 (예: 00435) — 서로 다른 코드 체계로 인해 클라이언트 선택 후 프로젝트 검색이 항상 0건이던 버그 수정.
- backend: client_code → client_name 변환 후 Azure cache 정규화 매칭 ((주) / 주식회사 / 공백 제거 후 부분 일치)
- backend: PG fallback 도 client_id FK 사용
- 검증: 01403 (삼성생명) → 5건, 02489 (삼성물산) → 4건, 05319 (삼성전자) → 11건 (이전 모두 0건)

추가: Azure 진행 LOS=10 프로젝트 8,079건 캐시 적재 (로컬 dev 환경에서 pymssql native binary 손상 + .env 자격증명 mock 처리 문제 발견·복구).

PR: https://github.com/ksyjerry/my_budget/pull/1 / Commit: 4a8209a (영역 1)
PR: https://github.com/ksyjerry/my_budget/pull/9 / Commit: 2dc48d1 (S8 후속)""")

# row 74 (#71) — Step 2 Excel export 화면 상이
add(74, EXCEL_REMOVED_NOTE + """

원래 답변 (영역 4): members/export 헤더 한글 표준화 + 팀 컬럼 추가.""")

# row 75 (#72) — Step 2 Excel 일부 컬럼만 업로드
add(75, EXCEL_REMOVED_NOTE + """

원래 답변 (영역 4): members/upload 사번+이름만 필수, 직급/팀 누락 시 마스터 lookup.""")

# row 77 (#74) — Step 3 저장 후 엑셀 다운로드 공란
add(77, EXCEL_REMOVED_NOTE + """

원래 답변 (영역 5): Step 3 export 12개월 모두 컬럼 + sort_order 정렬 + db.expire_all() 활성.""")

# row 80 (#77) — Appendix 다운로드 차단 경고
add(80, """🗑 Appendix 페이지 자체 제거됨 (S8 — 2026-04-29)

원래 fix (영역 6): 응답 헤더 (Content-Type + RFC 5987 filename*=UTF-8'') 보강.

PR: https://github.com/ksyjerry/my_budget/pull/9""")

# row 85 (#82) — Step 3 template 업로드 에러
add(85, EXCEL_REMOVED_NOTE + """

원래 답변 (영역 5): Excel template export/upload round-trip 일관성 + IP 노출 (#111) sanitize.""")

# row 89 (#86) — Step 2 Excel 업로드 오류 표시
add(89, EXCEL_REMOVED_NOTE + """

원래 답변 (영역 4): members/upload 응답 schema {imported, skipped, errors: [{row, col, error}]} — 행 단위 오류 누적.""")

# row 106 (#103) — Step 3 Template 다운로드
add(106, EXCEL_REMOVED_NOTE + """

원래 답변 (영역 5): Step 3 toolbar "📥 빈 Template 다운로드" 버튼 + #119 dropbox 설계.""")

# row 107 (#104) — 초기화 후 엑셀 다운로드 공란
add(107, EXCEL_REMOVED_NOTE + """

원래 답변 (영역 5): Excel export 가 항상 latest DB state 읽음 (db.expire_all()).""")

# row 108 (#105) — Excel 빈 월 누락
add(108, EXCEL_REMOVED_NOTE + """

원래 답변 (영역 5): Excel export 12개월 모두 고정 컬럼 + 0 인 월도 빈 셀.""")

# row 109 (#106) — Excel 업로드 후 비활성 사라짐
add(109, EXCEL_REMOVED_NOTE + """

원래 답변 (영역 5): Excel upload merge 방식 (truncate 아님) — 업로드에 없는 행은 enabled=false 유지.""")

# row 114 (#111) — 전체 V 토글 → S8에서 그리드 V 헤더 마스터 체크박스로 이동
add(114, """✅ 영역 5 fix + S8 후속 그리드 V 열 마스터 체크박스로 이동 완료

영역 5 (PR #5): Step 3 toolbar "전체 V 체크/해제" 토글 버튼 (1번 클릭 양방향).

S8 후속 fix (PR #9, 2026-04-29): 사용자 추가 요청 — 토글을 더 직관적으로 그리드 안 V 열 헤더로 이동.
- Toolbar의 "전체 V 해제/체크" button 제거
- MonthGrid V 열 thead 에 마스터 <input type="checkbox"> 배치
- 모두 enabled → checked / 일부만 enabled → indeterminate (ref 로 설정) / 모두 disabled → unchecked
- 클릭 시 모든 행의 enabled flag 일괄 토글

PR: https://github.com/ksyjerry/my_budget/pull/5 (영역 5)
PR: https://github.com/ksyjerry/my_budget/pull/9 / Commit: fa48734 (S8 후속)""")

# row 115 (#112) — 전체 V 해제
add(115, """✅ 영역 5 fix + S8 후속 마스터 체크박스로 통합 (#114 와 동일)

S8 후속: 단일 마스터 체크박스 (그리드 V 열 헤더) 가 양방향 처리 — 전체 체크 / 전체 해제 / indeterminate (일부) 3 상태.""")

# row 116 (#113) — Excel 업로드 후 단위 삭제
add(116, EXCEL_REMOVED_NOTE + """

원래 답변 (영역 5): Excel upload merge 방식 — 파일에 없는 budget_unit 은 enabled=false 비활성 유지.""")

# row 119 (#116) — 화면/엑셀 관리단위 순서 다름
add(119, EXCEL_REMOVED_NOTE + """

원래 답변 (영역 5): Export + UI 모두 budget_unit_master.sort_order 기준 정렬로 통일.

(엑셀이 제거되었지만 UI 정렬은 sort_order 기준으로 유지됨.)""")

# row 121 (#118) — 엑셀 템플릿 dropbox
add(121, EXCEL_REMOVED_NOTE + """

원래 답변 (영역 5): blank-export endpoint — openpyxl DataValidation + named range "BudgetUnitList".""")


# ─── 사용자 테스트 필요 항목 (v3 — Areas 1-7 + S8) ────────────────────────

USER_TEST_ITEMS = [
    # === Area 1 (회귀 7건) ===
    ("67", "영역 1", "배포 위생", "P0",
     "Docker prod build에서 dev overlay 0건 — /login, /overview-person, /budget-input 등 모든 페이지",
     "EL 계정",
     "nextjs-portal·data-nextjs-toast·data-nextjs-dialog 0개. console error 0",
     "PR #1"),

    ("68", "영역 1", "Step 1", "P1",
     "QRP 사번 입력 → 다른 필드 → QRP 다시 클릭 → 추가 입력",
     "EL 계정 (170661)",
     "값 유지 + 추가 입력 정상", "PR #1"),

    ("69", "영역 1", "Step 1", "P1",
     "신규 프로젝트 → 클라이언트 검색 skip → 프로젝트 검색만",
     "EL 또는 PM 계정",
     "Azure 진행 LOS=10 프로젝트 결과 표시 (로컬 dev 시 pymssql + .env 자격증명 필요)", "PR #1 + S8 후속 fix"),

    ("70", "영역 1", "Step 1", "P1",
     "총 계약시간 12345 입력 → readOnly 필드 표시",
     "EL 계정",
     "ET 잔여시간에 '12,345' 천단위 콤마", "PR #1"),

    ("71", "영역 1", "Step 2", "P1",
     "휴직/퇴사 사번 입력 → Enter",
     "EL 계정 + 휴직 시드 사번",
     "alert 표시 + 등록 차단 + input clear", "PR #1, INACTIVE_EMPNO 시드 필요"),

    ("74", "영역 1", "Step 1/3", "P1",
     "(a) AX/DX -1 (b) Step 3 cell 0.24 (c) 301",
     "EL 계정",
     "(a) 0 clamp (b) 0.25 snap (c) 300 clamp", "PR #1"),

    ("99", "영역 1", "Step 1", "P1",
     "Step 1 모든 button 비겹침 — width 1024/1280/1920",
     "EL 계정",
     "어떤 width 에서도 button 겹침 없음", "PR #1"),

    # === Area 2 ===
    ("79+82", "영역 2", "Budget 목록", "P0",
     "PM 계정으로 /budget-input 접근",
     "PM 계정 (본인이 PM인 프로젝트 보유)",
     "본인 PM 프로젝트 모두 표시 (이전: EL만 보였음)", "PR #2"),

    ("84", "영역 2", "Budget 목록", "P2",
     "/budget-input 화면 상단",
     "EL 계정",
     "'빈 Budget Template 다운로드' 버튼 미노출", "PR #2 + S8: blank-export 자체 제거"),

    ("85", "영역 2", "Budget 목록", "P2",
     "각 행에 '마지막 수정' 날짜 표시",
     "EL/PM 계정",
     "ko-KR 형식으로 표시", "PR #2"),

    ("120", "영역 2", "Budget 목록", "P1",
     "상태 필터 (전체/작성중/작성완료/승인완료) 선택",
     "EL 계정",
     "각 상태별 결과 즉시 변경", "PR #2"),

    ("121", "영역 2", "Budget 목록", "P1",
     "검색 'sk', 'SK', 'Sk텔레콤', 'SK텔레콤'",
     "EL 계정 (SK텔레콤 프로젝트 보유)",
     "모두 case-insensitive 매칭", "PR #2"),

    ("61/98", "영역 2", "POL-04 워크플로우", "P0",
     "PM submit → EL approve → unlock 전체 흐름 + Staff 직접 API 호출",
     "PM (170661) + EL + Staff",
     "각 단계 상태 변경 + Staff 차단", "PR #2"),

    ("64", "영역 2", "POL-05 daily sync", "P1",
     "Backend 재시작 → APScheduler 'daily_tba_sync' 등록 + 04:00 KST 자동 실행",
     "admin 계정",
     "Backend log 'Scheduler started' + 익일 신규 TBA 자동 추가", "PR #2"),

    # === Area 3 ===
    ("57", "영역 3", "Step 1", "P1",
     "클라이언트 A → B → C 변경 시퀀스",
     "EL 계정",
     "매번 의존 필드 (산업분류·자산규모) 새 client로 갱신", "PR #3"),

    ("62 + S8 후속", "영역 3 + S8", "Step 1", "P1",
     "Step 1 화면 진입 — 섹션 순서 확인",
     "EL 계정",
     "서비스 분류 → 클라이언트 기본정보 → 프로젝트 정보 → 시간 배분 순서로 위→아래 배치",
     "PR #3 + PR #9 (2dc48d1)"),

    ("86/100", "영역 3", "Step 1", "P1",
     "더본코리아 / 롯데지알에스 '이전 프로젝트 정보 가져오기'",
     "EL 계정 (시드 프로젝트 필요)",
     "alert + 시간/구성원/template 자동 채움", "PR #3, CLONE_SOURCE_PROJECT 시드 필요"),

    ("101", "영역 3", "Step 1", "P1",
     "Step 1 화면 — Fulcrum/RA-Staff/Specialist 라벨 0개. Step 3에서는 여전히 존재",
     "EL 계정",
     "Step 1 입력칸 제거 (POL-06 (a))", "PR #3"),

    # === Area 4 (S8 에서 Excel 부분 deprecated) ===
    ("88", "영역 4", "Step 2", "P1",
     "EmployeeSearch 결과 dropdown",
     "EL 계정",
     "팀명 column 표시 (동명이인 구분)", "PR #4"),

    ("102", "영역 4", "Step 2", "P2",
     "'+ TBD' / '+ NS' / '+ Associate' 버튼 클릭",
     "EL 계정",
     "placeholder 행 추가됨", "PR #4"),

    ("103", "영역 4", "Step 2", "P2",
     "Fulcrum/RA-Staff/Specialist 입력",
     "EL 계정",
     "<select> dropdown — 자유 텍스트 불가", "PR #4"),

    # === Area 5 (S8 에서 Excel 부분 deprecated, UI 부분 유지) ===
    ("Group B+G", "영역 5", "초기화 + 정책", "P1",
     "초기화 → step 이동 → 재진입 깨끗 + 종료월 입력 → 저장 → 재로드 유지",
     "EL 계정",
     "frontend+backend reset 일관 + fiscal_end 유지", "PR #5"),

    ("Group D", "영역 5", "Step 3 layout", "P2",
     "Step 3 합계 칸 정렬 확인",
     "EL 계정",
     "정상 정렬 (col 5 합계 값) + S8 sticky 적용 후 스크롤해도 정렬 유지", "PR #5 + PR #9"),

    ("Group E", "영역 5", "AI Assist", "P1",
     "비감사 service_type 추천 / 계속감사 → 초도감사 0 / cancel handling / 등록오류 IP 미노출",
     "EL 계정 (LLM API 작동 환경)",
     "service_type 매핑 + 컨텍스트 정확", "PR #5"),

    ("Group F", "영역 5", "금융업 시드 + POL-03", "P0",
     "감사 프로젝트 → Step 3 에 금융업 78행 budget_unit 노출",
     "EL 계정",
     "대출채권/보험계약부채/KICS 등 표시 + subcategory_name 분리", "PR #5"),

    # === Area 6 ===
    ("POL-01 (b)", "영역 6", "Budget 정의", "P0",
     "Overview KPI / Project 테이블 / Tracking / Summary 모두 Budget = 총계약 - AX/DX 확인",
     "EL 계정",
     "모든 view 동일 정의 + 실질 Progress 컬럼", "PR #6"),

    ("65", "영역 6", "Details", "P2",
     "타 LoS 인원 — 이름 미등록 시 fallback 표시",
     "EL 계정 (시드 필요)",
     "'이름 미등록 ({empno})' 표시", "PR #6"),

    ("77", "영역 6", "분류", "P2",
     "service_type 표시 — '감사/비감사' 통일",
     "EL 계정",
     "기타 service_type 도 '비감사' 그룹", "PR #6"),

    ("93", "영역 6", "Overview", "P1",
     "QRP empno 있는 프로젝트 — Actual 표시",
     "EL 계정 (시드 + Azure SQL 가용)",
     "QRP TMS 조회 정상", "PR #6"),

    ("94", "영역 6", "Overview", "P1",
     "Budget 없는 인원 actuals — Overview Project 테이블",
     "EL 계정 + TMS data",
     "actuals 집계 포함", "PR #6"),

    ("95", "영역 6", "Overview filter", "P1",
     "연월 dropdown — 4월~3월 12개월",
     "EL 계정",
     "월별 선택 가능 + 누적 옵션", "PR #6"),

    ("96", "영역 6", "Overview filter", "P1",
     "프로젝트 선택 → EL/PM dropdown",
     "EL 계정",
     "해당 프로젝트의 EL/PM 만 활성화 (cascading)", "PR #6"),

    ("POL-08 (b)", "영역 6", "Budget Tracking 권한", "P0",
     "EL/admin/PM/Staff 각각 Tracking 화면 접근",
     "각 페르소나 계정",
     "EL+admin OK / PM+Staff 403", "PR #6"),

    # === Area 7 (Wizard 분해) ===
    ("Area 7 분해", "영역 7", "Wizard 구조", "P1",
     "Step 1/2/3 wizard 흐름 — 신규 프로젝트 생성 → 작성완료 → 승인 전체 시나리오",
     "EL + PM",
     "분해 전과 동일 동작 (3,150 LOC → 449 LOC + 12 파일)", "PR #8"),

    # === S8 (Excel 제거 + Grid UX) ===
    ("S8 — Excel 제거", "S8", "메뉴/버튼", "P0",
     "Step 2/3 화면 + Appendix 페이지 + 사이드바 메뉴 점검",
     "EL 계정",
     "Excel 다운/업로드 버튼 0개 + Appendix 메뉴 미노출 + /appendix 직접 접근 시 404",
     "PR #9"),

    ("S8 — Step 3 Sticky", "S8", "Grid UX", "P0",
     "Step 3 그리드 스크롤 — 가로/세로 모두",
     "EL 계정 (Budget 입력 진행 중 프로젝트)",
     "헤더 / 좌측 4 column (V/대분류/관리단위/담당자) / 우측 합계 column 모두 sticky 유지",
     "PR #9 (9c34f41)"),

    ("S8 — 분배 도우미", "S8", "Grid UX", "P1",
     "Step 3 toolbar '📊 분배 도우미' 클릭 → 3가지 모드",
     "EL 계정",
     "균등 / 기말 집중 / 유사회사 비율 — 각 모드별 0.25h 단위 분배 + ET 잔여시간 일치",
     "PR #9 (a60283a)"),

    ("S8 — 실시간 검증", "S8", "Grid UX", "P1",
     "Step 3 행에 담당자 미지정 / 시간 미입력 상태",
     "EL 계정",
     "담당자 cell 빨간 ring + 합계 cell 노란 ring + SummaryRow 진행률 bar + 일치/초과/부족 표시",
     "PR #9 (5e58081)"),

    ("S8 — 검색·접기", "S8", "Grid UX", "P1",
     "Step 3 검색 input 입력 + 대분류 첫 행 클릭으로 접기/펼치기",
     "EL 계정",
     "대분류/관리단위/담당자/사번 매칭 + ▶/▼ 토글 + 새로고침 후 sessionStorage 복원",
     "PR #9 (d7eb253)"),

    ("S8 후속 — Step 1 재배치", "S8", "Step 1 layout", "P1",
     "Step 1 화면 진입",
     "EL 계정",
     "서비스 분류 → 클라이언트 기본정보 → 프로젝트 정보 → 시간 배분 순서",
     "PR #9 (2dc48d1)"),

    ("S8 후속 — 프로젝트 검색", "S8", "Step 1 search", "P0",
     "클라이언트 검색 (예: 삼성생명 01403) → 프로젝트 검색",
     "EL 계정 (Azure 가용 환경)",
     "해당 클라이언트의 Azure 진행 프로젝트 모두 표시 (이전: 0건)",
     "PR #9 (2dc48d1) — client_code → client_name 매핑 fix"),

    ("S8 후속 — V 토글", "S8", "Step 3 UX", "P2",
     "Step 3 그리드 V 열 헤더 마스터 체크박스 클릭",
     "EL 계정",
     "전체 행 enabled toggle / 일부 enabled 시 indeterminate 표시",
     "PR #9 (fa48734)"),

    ("S8 진단 — Azure 연결", "S8", "로컬 dev 환경", "P2",
     "로컬 백엔드 시작 → Azure project cache 적재 확인",
     "개발자 (로컬 dev)",
     "Backend log 'Azure project cache loaded: 8079 projects' (이전: pymssql 손상 + .env mock 문제)",
     "PR #9 — pymssql 재컴파일 + .env 자격증명 main worktree 에서 복사"),

    # === 공통 ===
    ("CI 게이트", "영역 1+2 공통", "CI 게이트", "P0",
     "(a) <input type=number> 추가 PR (b) 게이트 미통과 PR merge 시도",
     "GitHub repo write 권한",
     "(a) Grep Guards 실패 (b) merge 차단", "Branch protection 적용 후"),

    ("Layer 3 — 모든 Area + S8", "영역 1-8 종료", "사용자 컨펌", "P0",
     "각 Area 의 qa-checklist 일괄 staging 검증 (S8 manual QA 포함)",
     "EL + PM + Staff + admin",
     "모든 결함 차단/fix 유지 + qa-checklist all PASS",
     "S8 sign-off → main merge"),

    # === POL 외부 결정자 컨펌 ===
    ("POL-01", "정책", "외부 결정자", "P0",
     "Budget 정의 (b) 총계약-AX/DX confirm",
     "김동환 (4/27 회의)",
     "정식 확정 → policy-decisions.md 업데이트", "영역 6 적용"),

    ("POL-02", "정책", "외부 결정자", "P0",
     "통상자문/내부회계 (b) service_type별 다름",
     "신승엽 (4/27)",
     "정식 확정 + 통상자문 UI mini-cycle 트리거", "영역 5 schema 준비됨"),

    ("POL-03", "정책", "외부 결정자", "P0",
     "비감사 관리단위 (a) 소분류명 별도 컬럼",
     "나형우/김지민 (4/27)",
     "정식 확정 → 영역 5 시드 패턴 유지", "alembic 007 적용됨"),

    ("POL-04", "정책", "외부 결정자 (provisional)", "P1",
     "EL 승인 워크플로우 (b) 표준형",
     "김동환",
     "정식 확정 OR (a)/(c) 결정 시 영역 2 후속 fix", "영역 2 적용"),

    ("POL-05", "정책", "외부 결정자 (provisional)", "P2",
     "TBA sync (d) 하이브리드 (Daily 04:00 + manual)",
     "김미진",
     "정식 확정 OR 다른 안 시 영역 2 후속", "영역 2 적용"),

    ("POL-06", "정책", "외부 결정자 (provisional)", "P0",
     "RA 주관 FLDT (a) Step 1 제거 / Step 3에만",
     "홍상호",
     "정식 확정 OR 다른 안 시 영역 3 후속", "영역 3 적용"),

    ("POL-07", "정책", "외부 결정자 (provisional)", "P1",
     "프로젝트 기간 (c) 시작 자동 + 끝 수동",
     "신승엽",
     "정식 확정 OR 다른 안 시 영역 5 후속", "alembic 007 + 영역 5 적용"),

    ("POL-08", "정책", "외부 결정자 (provisional)", "P1",
     "Budget Tracking 권한 (b) EL+admin",
     "김동환",
     "정식 확정 OR PM 노출 결정 시 영역 6 후속", "영역 6 적용"),

    ("POL-09", "정책", "내부 결정", "P3",
     "TBA sync 권한 admin only",
     "owner",
     "admin only 권장 (YAGNI)", "영역 2 회고 포함"),
]


def fill_answers(ws):
    align = Alignment(wrap_text=True, vertical="top")
    border = Side(style="thin", color="DDDDDD")
    cell_border = Border(top=border, bottom=border, left=border, right=border)
    for row, body in ANSWERS.items():
        c = ws.cell(row=row, column=7, value=body)
        c.alignment = align
        c.border = cell_border
        ws.row_dimensions[row].height = max(60, min(400, body.count("\n") * 14 + 40))
    ws.column_dimensions["G"].width = 80


def add_user_test_sheet(wb):
    if "사용자_테스트_필요_항목" in wb.sheetnames:
        del wb["사용자_테스트_필요_항목"]
    ws = wb.create_sheet("사용자_테스트_필요_항목", 1)

    headers = ["피드백 No", "영역", "카테고리", "우선순위",
               "테스트 시나리오", "사용 계정·데이터", "기대 결과", "비고"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="D04A02")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Side(style="thin", color="888888")
    cell_border = Border(top=border, bottom=border, left=border, right=border)

    ws.cell(row=1, column=1, value=(
        f"My Budget+ 2.0 — 사용자 테스트 필요 항목 (S7 + S8 완료, v3 작성: {datetime.now().strftime('%Y-%m-%d')}) | "
        f"S7 6 영역 + S8 (Excel 제거 + Grid UX) + S8 후속 fixes (Step 1 재배치 / 프로젝트 검색 / V 토글)"
    ))
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.row_dimensions[1].height = 36

    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = cell_border
    ws.row_dimensions[3].height = 32

    align_top = Alignment(wrap_text=True, vertical="top")
    align_top_center = Alignment(wrap_text=True, vertical="top", horizontal="center")
    priority_color = {
        "P0": PatternFill("solid", fgColor="FFCDD2"),
        "P1": PatternFill("solid", fgColor="FFE0B2"),
        "P2": PatternFill("solid", fgColor="FFF9C4"),
        "P3": PatternFill("solid", fgColor="C8E6C9"),
    }
    for r_idx, item in enumerate(USER_TEST_ITEMS, 4):
        for c_idx, val in enumerate(item, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.border = cell_border
            if c_idx == 4:
                c.alignment = align_top_center
                c.fill = priority_color.get(val, PatternFill())
            else:
                c.alignment = align_top
        max_len = max(len(str(v)) for v in item)
        ws.row_dimensions[r_idx].height = max(40, min(180, (max_len // 50 + 1) * 28))

    widths = {"A": 22, "B": 14, "C": 22, "D": 10, "E": 50, "F": 30, "G": 50, "H": 30}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A4"


def main():
    if not SRC.exists():
        raise SystemExit(f"Source file missing: {SRC}")
    shutil.copy(SRC, DST)
    wb = openpyxl.load_workbook(DST)

    if "#01_사용자 feedback" not in wb.sheetnames:
        raise SystemExit("Sheet '#01_사용자 feedback' missing")

    fill_answers(wb["#01_사용자 feedback"])
    add_user_test_sheet(wb)

    wb.save(DST)
    print(f"Updated: {DST}")
    print(f"  답변 작성: {len(ANSWERS)} 행")
    print(f"  신규 시트: 사용자_테스트_필요_항목 ({len(USER_TEST_ITEMS)} 항목)")


if __name__ == "__main__":
    main()
